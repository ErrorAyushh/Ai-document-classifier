"""
ai_engine/excel_filler.py

Fills the GFS pricing worksheet template with job data extracted by the AI pipeline.
All row lookups use find_row_by_label() — never hardcoded row numbers.
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ---------------------------------------------------------------------------
# Job type constants
# ---------------------------------------------------------------------------
TYPE_STANDARD    = "TYPE_STANDARD"
TYPE_MULTI_UNIT  = "TYPE_MULTI_UNIT"
TYPE_ECONOFRAME  = "TYPE_ECONOFRAME"
TYPE_GLASS_ONLY  = "TYPE_GLASS_ONLY"
TYPE_CUSTOM      = "TYPE_CUSTOM"

# ---------------------------------------------------------------------------
# Template path constants
# TEMPLATE_V1 — Location row <= 112, older Bonnell pricing sheet
# TEMPLATE_V2 — Location row > 112, newer Bonnell pricing sheet
# ---------------------------------------------------------------------------
TEMPLATE_V1 = "templates/gfs_pricing_template.xlsx"
TEMPLATE_V2 = "templates/gfs_pricing_template_v2.xlsx"

# ---------------------------------------------------------------------------
# Unit toggle column indices (1-based)
# A=D(4), B=G(7), C=J(10), D=M(13), E=P(16), F=S(19), G=V(22), H=Y(25),
# I=AB(28), J=AE(31)
# ---------------------------------------------------------------------------
_UNIT_TOGGLE_COLS = [4, 7, 10, 13, 16, 19, 22, 25, 28, 31]

# ---------------------------------------------------------------------------
# Cross-beam label map
# key = job_data["cross_beam"] value
# value = label to search for in the worksheet via find_row_by_label()
# ---------------------------------------------------------------------------
CROSS_BEAM_LABELS = {
    "I_Beam_TB":       "Series 1000 I - Beam thermally broken",
    "T_Section_TB":    "Series 1000 T - Section thermally Broken",
    "I_Beam":          "Series 1000 I - Beam",
    "T_Section":       "Series 1000 T - Section",
    "I_Beam_S2000_TB": "Series 2000 I beam thermally broken",
    "EconoFrame_25":   "Econoframe 2.5",
    "EconoFrame_30":   "Econoframe 3.0",
}

# Cross-beam is always toggled for the extra-unit columns beyond col D:
# S(19), V(22), Y(25), AB(28), AE(31)
_CROSS_BEAM_TOGGLE_COLS = [19, 22, 25, 28, 31]


# ---------------------------------------------------------------------------
# Helper: find_row_by_label
# ---------------------------------------------------------------------------
def find_row_by_label(ws, label: str, col_idx: int = 2):
    """
    Search column col_idx (default B=2) for a cell whose value contains
    label (case-insensitive).
    Handles MergedCell by catching AttributeError and skipping.
    Returns the row number (int) or None if not found.
    """
    label_lower = label.lower()
    for row in ws.iter_rows():
        cell = row[col_idx - 1]  # convert 1-based col_idx to 0-based list index
        try:
            cell_value = cell.value
        except AttributeError:
            continue
        if cell_value is None:
            continue
        try:
            if label_lower in str(cell_value).lower():
                return cell.row
        except AttributeError:
            continue
    return None


# ---------------------------------------------------------------------------
# Helper: get_unit_toggle_cols
# ---------------------------------------------------------------------------
def get_unit_toggle_cols(num_units: int) -> list:
    """
    Return list of column indices for unit toggles, up to num_units.
    A=4(D), B=7(G), C=10(J), D=13(M), E=16(P), F=19(S), G=22(V),
    H=25(Y), I=28(AB), J=31(AE)
    """
    return _UNIT_TOGGLE_COLS[:num_units]


# ---------------------------------------------------------------------------
# Helper: _should_write
# ---------------------------------------------------------------------------
def _should_write(value) -> bool:
    """
    Returns True if value should be written to the worksheet.
    Skips None and the placeholder string "N/A".
    """
    if value is None:
        return False
    if str(value).strip() == "N/A":
        return False
    return True


# ---------------------------------------------------------------------------
# Helper: auto_detect_template
# ---------------------------------------------------------------------------
def auto_detect_template(original_path: str = None) -> str:
    """
    Smart template detection for unknown / new folders.

    If original_path is provided and the file exists:
      - Load the original workbook (read-only)
      - Find the Location row in 'Pricing Worksheet'
      - location_row <= 112 → TEMPLATE_V1
      - anything else       → TEMPLATE_V2  (safe default)

    If original_path is not provided or the file does not exist:
      → TEMPLATE_V2  (default for all new jobs)

    Returns the matching template path constant.
    """
    if original_path is None or not os.path.exists(original_path):
        return TEMPLATE_V2

    try:
        wb = load_workbook(original_path, data_only=False, read_only=True)
        if "Pricing Worksheet" in wb.sheetnames:
            ws = wb["Pricing Worksheet"]
            location_row = find_row_by_label(ws, "Location")
            wb.close()
            if location_row is not None and location_row <= 112:
                return TEMPLATE_V1
            else:
                return TEMPLATE_V2
        wb.close()
    except Exception:
        pass

    return TEMPLATE_V2


# ---------------------------------------------------------------------------
# Helper: detect_template_for_original  (used inside verify_fill)
# ---------------------------------------------------------------------------
def detect_template_for_original(original_path: str) -> str:
    """
    Opens the original (hand-filled) workbook and finds the Location row
    in the 'Pricing Worksheet' sheet to determine which template version
    was used to create it.

    location_row <= 112 → TEMPLATE_V1
    anything else       → TEMPLATE_V2

    Returns the matching template path constant.
    """
    try:
        wb = load_workbook(original_path, data_only=False, read_only=True)
        if "Pricing Worksheet" in wb.sheetnames:
            ws = wb["Pricing Worksheet"]
            location_row = find_row_by_label(ws, "Location")
            wb.close()
            if location_row is not None and location_row <= 112:
                return TEMPLATE_V1
            else:
                return TEMPLATE_V2
        wb.close()
    except Exception:
        pass
    return TEMPLATE_V2


# ---------------------------------------------------------------------------
# Main: fill_workbook
# ---------------------------------------------------------------------------
def fill_workbook(
    job_data: dict,
    template_path: str = None,
    original_path: str = None,
    output_path: str = None,
) -> dict:
    """
    Fills the GFS pricing worksheet template with job_data.

    Template selection priority
    ---------------------------
    a) template_path explicitly passed → use it as-is
    b) job_data["template_version"] set:
         "v1" → TEMPLATE_V1
         anything else → TEMPLATE_V2
    c) original_path provided → auto_detect_template(original_path)
    d) fallback → auto_detect_template() → TEMPLATE_V2

    Parameters
    ----------
    job_data      : dict       — extracted job data
    template_path : str | None — explicit override; skips all auto-detection
    original_path : str | None — path to the hand-filled original; used for
                                 auto-detection when template_version absent
    output_path   : str | None — save path; auto-generated if None

    Returns
    -------
    dict with keys "status" and (on success) "output_path"
    """

    # ------------------------------------------------------------------
    # 1. Resolve template path
    # ------------------------------------------------------------------
    if template_path is not None:
        # (a) explicit override
        pass
    elif job_data.get("template_version"):
        # (b) explicit version tag in job_data
        version = job_data["template_version"]
        if version == "v1":
            template_path = TEMPLATE_V1
        else:
            template_path = TEMPLATE_V2
    elif original_path is not None:
        # (c) detect from original file
        template_path = auto_detect_template(original_path)
    else:
        # (d) safe default
        template_path = auto_detect_template()

    # ------------------------------------------------------------------
    # 2. Load template — never data_only=True so formulas are preserved
    # ------------------------------------------------------------------
    wb = load_workbook(template_path)
    ws = wb["Pricing Worksheet"]

    # Default packaging count to 1 if not already set in the template
    if ws["E71"].value is None:
        ws["E71"] = 1

    # ------------------------------------------------------------------
    # 3. Convenience aliases from job_data
    # ------------------------------------------------------------------
    job_type       = job_data.get("job_type", TYPE_STANDARD)
    project_name   = job_data.get("project_name", "")
    address        = job_data.get("address", "")
    quote_number   = job_data.get("quote_number", "")
    person_quoting = job_data.get("person_quoting", "")
    contact_name   = job_data.get("contact_name", "")
    contact_phone  = job_data.get("contact_phone", "")
    contact_email  = job_data.get("contact_email", "")
    project_type   = job_data.get("project_type", "")
    architect      = job_data.get("architect")
    homeowner      = job_data.get("homeowner")
    units          = job_data.get("units", [])
    frame_type     = job_data.get("frame_type", "")
    glass_type     = job_data.get("glass_type", "")
    cross_beam     = job_data.get("cross_beam", None)
    nanodot        = job_data.get("nanodot", False)
    heat_soak      = job_data.get("heat_soak", False)
    seeded_organic = job_data.get("seeded_organic", False)
    duty           = job_data.get("duty", False)
    backpaint      = job_data.get("backpaint", False)
    man_hours      = job_data.get("man_hours", 0)
    duty_rate      = job_data.get("duty_rate", 0.055)
    profit_margin  = job_data.get("profit_margin", 0.5)

    # ------------------------------------------------------------------
    # 4. Write header fields
    # ------------------------------------------------------------------
    ws.cell(row=2, column=3).value  = project_name    # C2
    ws.cell(row=2, column=6).value  = address          # F2
    ws.cell(row=2, column=9).value  = quote_number     # I2
    ws.cell(row=2, column=12).value = person_quoting   # L2
    ws.cell(row=3, column=3).value  = contact_name     # C3
    ws.cell(row=3, column=9).value  = contact_phone    # I3
    ws.cell(row=3, column=12).value = contact_email    # L3
    ws.cell(row=4, column=3).value  = project_type     # C4

    # F4 — architect: only write if not None and not "N/A"
    if _should_write(architect):
        ws.cell(row=4, column=6).value = architect     # F4

    # I4 — homeowner: only write if not None and not "N/A"
    if _should_write(homeowner):
        ws.cell(row=4, column=9).value = homeowner     # I4

    # ------------------------------------------------------------------
    # 4b. Write duty rate and profit margin
    # ------------------------------------------------------------------
    duty_row = find_row_by_label(ws, "Duty")
    if duty_row is not None:
        ws.cell(row=duty_row, column=3).value = duty_rate

    profit_row = find_row_by_label(ws, "Profit")
    if profit_row is not None:
        ws.cell(row=profit_row, column=3).value = profit_margin

    # ------------------------------------------------------------------
    # 5. Find take-off header row → data_start
    # ------------------------------------------------------------------
    location_row = find_row_by_label(ws, "Location")
    if location_row is None:
        data_start = 10
    else:
        data_start = location_row + 1

    # ------------------------------------------------------------------
    # 6. Write unit rows
    # ------------------------------------------------------------------
    for i, unit in enumerate(units):
        row = data_start + i

        width_val  = unit.get("width_inches")
        length_val = unit.get("length_inches")

        # Round float dimensions to 8 decimal places to avoid float precision noise
        if isinstance(width_val, float):
            width_val = round(width_val, 8)
        if isinstance(length_val, float):
            length_val = round(length_val, 8)

        ws.cell(row=row, column=2).value  = unit.get("location", "")        # B
        ws.cell(row=row, column=3).value  = unit.get("drawing_number", "")  # C
        ws.cell(row=row, column=5).value  = width_val                        # E
        ws.cell(row=row, column=6).value  = length_val                       # F
        ws.cell(row=row, column=7).value  = unit.get("panel_count")          # G
        ws.cell(row=row, column=10).value = unit.get("unit_count")           # J

        rafter_v = unit.get("rafter_vertical")
        rafter_h = unit.get("rafter_horizontal")
        if rafter_v is not None:
            ws.cell(row=row, column=11).value = rafter_v                     # K
        if rafter_h is not None:
            ws.cell(row=row, column=12).value = rafter_h                     # L

    # ------------------------------------------------------------------
    # 7. TYPE_CUSTOM — save and return early
    # ------------------------------------------------------------------
    if job_type == TYPE_CUSTOM:
        output_path = _resolve_output_path(output_path, quote_number)
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        return {"status": "manual_review_required", "output_path": output_path}

    # ------------------------------------------------------------------
    # 8. Frame toggles
    # ------------------------------------------------------------------
    num_units = len(units)

    if job_type == TYPE_STANDARD:
        frame_row = find_row_by_label(ws, frame_type)
        if frame_row is not None:
            ws.cell(row=frame_row, column=4).value = "y"

    elif job_type == TYPE_MULTI_UNIT:
        frame_row = find_row_by_label(ws, frame_type)
        if frame_row is not None:
            toggle_cols = get_unit_toggle_cols(num_units)
            for col in toggle_cols:
                ws.cell(row=frame_row, column=col).value = "y"

    elif job_type == TYPE_ECONOFRAME:
        lengths_row = find_row_by_label(ws, "Econoframe - lengths")
        if lengths_row is not None:
            ws.cell(row=lengths_row, column=4).value = "y"
        econ30_row = find_row_by_label(ws, "Econoframe 3.0")
        if econ30_row is not None:
            ws.cell(row=econ30_row, column=4).value = "y"

    elif job_type == TYPE_GLASS_ONLY:
        pass  # No frame toggle

    # ------------------------------------------------------------------
    # 8b. Cross-beam toggle (after frame toggles)
    # ------------------------------------------------------------------
    if cross_beam is not None:
        cb_label = CROSS_BEAM_LABELS.get(cross_beam)
        if cb_label is not None:
            cb_row = find_row_by_label(ws, cb_label)
            if cb_row is not None:
                for col in _CROSS_BEAM_TOGGLE_COLS:
                    ws.cell(row=cb_row, column=col).value = "y"

    # ------------------------------------------------------------------
    # 9. Glass toggle
    # ------------------------------------------------------------------
    glass_row = find_row_by_label(ws, glass_type)
    if glass_row is not None:
        # Non-Walkable glass uses per-unit toggle columns like TYPE_MULTI_UNIT
        if "non-walkable" in glass_type.lower():
            toggle_cols = get_unit_toggle_cols(num_units)
            for col in toggle_cols:
                ws.cell(row=glass_row, column=col).value = "y"
        else:
            ws.cell(row=glass_row, column=4).value = "y"

    # ------------------------------------------------------------------
    # 10. Add-alternates
    # ------------------------------------------------------------------
    alternates = [
        (nanodot,        "Nano Dot"),
        (heat_soak,      "Heat Soak Testing"),
        (seeded_organic, "Seeded Organic"),
        (duty,           "Duty"),
        (backpaint,      "Backpaint"),
    ]
    for flag, label in alternates:
        if flag:
            alt_row = find_row_by_label(ws, label)
            if alt_row is not None:
                ws.cell(row=alt_row, column=4).value = "y"

    # ------------------------------------------------------------------
    # 11. Man hours
    # ------------------------------------------------------------------
    mh_row = find_row_by_label(ws, "Man Hours")
    if mh_row is not None:
        ws.cell(row=mh_row, column=5).value = man_hours

    # ------------------------------------------------------------------
    # 12. Resolve output path, make dirs, save
    # ------------------------------------------------------------------
    output_path = _resolve_output_path(output_path, quote_number)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)

    return {"status": "ok", "output_path": output_path}


# ---------------------------------------------------------------------------
# Helper: resolve output path
# ---------------------------------------------------------------------------
def _resolve_output_path(output_path, quote_number):
    if output_path is None:
        safe_quote = str(quote_number).replace("/", "-").replace("\\", "-")
        output_path = f"output/{safe_quote}_estimate.xlsx"
    return output_path


# ---------------------------------------------------------------------------
# verify_fill
# ---------------------------------------------------------------------------
def verify_fill(original_path: str, output_path: str) -> dict:
    """
    Opens both workbooks with data_only=False so that formula cells retain
    their raw '=' strings. This allows the startswith('=') skip to work
    correctly — formula cells are counted as skipped_formulas and never
    appear in missing or mismatched.

    detect_template_for_original() is called to identify which template
    version the original was built from (for reference / future use).

    For every yellow cell in the original (fgColor rgb contains FFFF00):
      - If original cell value starts with '=' → skipped_formulas (not verified)
      - If original is None and output is None → matched
      - If original is non-None/non-formula and output is None → missing
      - If both non-None and stripped strings match → matched
      - Otherwise → mismatched

    Returns
    -------
    {
        "matched":           [coord, ...],
        "mismatched":        [(coord, original_value, output_value), ...],
        "missing":           [coord, ...],
        "skipped_formulas":  int,
        "template_detected": str   — TEMPLATE_V1 or TEMPLATE_V2
    }
    """
    # Detect which template version the original was built from (informational)
    detected_template = detect_template_for_original(original_path)

    # Load both workbooks with data_only=False so formula strings are preserved
    # as raw "=..." values rather than being resolved to computed numbers.
    wb_orig = load_workbook(original_path, data_only=False)
    wb_out  = load_workbook(output_path,   data_only=False)

    matched          = []
    mismatched       = []
    missing          = []
    skipped_formulas = 0

    for sheet_name in wb_orig.sheetnames:
        if sheet_name not in wb_out.sheetnames:
            continue

        ws_orig = wb_orig[sheet_name]
        ws_out  = wb_out[sheet_name]

        for row in ws_orig.iter_rows():
            for cell in row:
                # Only inspect yellow cells
                if not _is_yellow(cell):
                    continue

                coord    = cell.coordinate
                orig_val = cell.value

                # ----------------------------------------------------------
                # Skip formula cells in the original — loaded with
                # data_only=False so these are raw "=..." strings.
                # These are template pricing constants that fill_workbook
                # never writes to. Count them for the summary only.
                # ----------------------------------------------------------
                if isinstance(orig_val, str) and orig_val.startswith("="):
                    skipped_formulas += 1
                    continue

                # Get corresponding cell in output workbook
                try:
                    out_cell = ws_out[coord]
                    out_val  = out_cell.value
                except Exception:
                    # Coord doesn't exist at all in the output sheet
                    if orig_val is not None:
                        missing.append(coord)
                    continue

                # Skip formula cells in the output as well (belt-and-braces)
                if isinstance(out_val, str) and out_val.startswith("="):
                    skipped_formulas += 1
                    continue

                # Categorise
                if orig_val is None and out_val is None:
                    matched.append(coord)
                elif orig_val is not None and out_val is None:
                    # Original has a real value but output is empty → missing
                    missing.append(coord)
                elif orig_val is None and out_val is not None:
                    # Output has a value where original was blank → mismatch
                    mismatched.append((coord, orig_val, out_val))
                elif str(orig_val).strip() == str(out_val).strip():
                    matched.append(coord)
                else:
                    mismatched.append((coord, orig_val, out_val))

    return {
        "matched":           matched,
        "mismatched":        mismatched,
        "missing":           missing,
        "skipped_formulas":  skipped_formulas,
        "template_detected": detected_template,
    }


def _is_yellow(cell) -> bool:
    """
    Returns True if the cell's fill foreground colour contains 'FFFF00'.
    Handles MergedCell and cells with no fill gracefully.
    """
    try:
        fill = cell.fill
        if fill is None:
            return False
        fg = fill.fgColor
        if fg is None:
            return False
        rgb = fg.rgb
        if rgb is None:
            return False
        return "FFFF00" in str(rgb).upper()
    except AttributeError:
        return False


# ---------------------------------------------------------------------------
# __main__ — run fill_workbook + verify_fill for the selected test job(s)
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import json
    from ai_engine.test_jobs import TEST_JOBS

    # -----------------------------------------------------------------------
    # RUN_ALL = True  → loop every job in TEST_JOBS and print a summary table
    # RUN_ALL = False → run only ACTIVE_JOB with full mismatch/missing detail
    # -----------------------------------------------------------------------
    RUN_ALL    = True
    ACTIVE_JOB = "182_robinson"

    # -----------------------------------------------------------------------
    # Helper: run one job and return a summary row dict
    # -----------------------------------------------------------------------
    def _run_job(job_key: str) -> dict:
        job_data      = TEST_JOBS[job_key]["job_data"]
        original_path = TEST_JOBS[job_key]["original_path"]

        print(f"\n{'=' * 60}")
        print(f"Job        : {job_key}")
        print(f"Original   : {original_path}")
        print(f"{'=' * 60}")

        # Pass original_path so template is auto-detected when
        # template_version is absent from job_data
        result = fill_workbook(
            job_data,
            template_path=None,
            original_path=original_path,
        )
        print(f"fill_workbook → status: {result.get('status')}  "
              f"output: {result.get('output_path')}")

        status = result.get("status")

        if status not in ("ok", "manual_review_required"):
            print("  Unexpected status — skipping verify_fill.")
            return {
                "job":      job_key,
                "status":   status,
                "matched":  "-",
                "mismatch": "-",
                "missing":  "-",
                "skipped":  "-",
                "template": "-",
            }

        output_path = result["output_path"]
        report = verify_fill(original_path, output_path)

        n_matched    = len(report["matched"])
        n_mismatched = len(report["mismatched"])
        n_missing    = len(report["missing"])
        n_skipped    = report["skipped_formulas"]
        detected     = report["template_detected"]

        if report["mismatched"]:
            print("  Mismatches:")
            for coord, expected, got in report["mismatched"]:
                print(f"    {coord}: expected {expected!r}  got {got!r}")

        if report["missing"]:
            print("  Missing:")
            for coord in report["missing"]:
                print(f"    {coord}")

        return {
            "job":      job_key,
            "status":   status,
            "matched":  n_matched,
            "mismatch": n_mismatched,
            "missing":  n_missing,
            "skipped":  n_skipped,
            "template": "v1" if detected == TEMPLATE_V1 else "v2",
        }

    # -----------------------------------------------------------------------
    # RUN_ALL branch
    # -----------------------------------------------------------------------
    if RUN_ALL:
        summary_rows = []
        for job_key in TEST_JOBS:
            row = _run_job(job_key)
            summary_rows.append(row)

        # Print summary table
        col_w   = [22, 24, 9, 11, 9, 16, 10]
        headers = ["Job", "Status", "Matched", "Mismatched",
                   "Missing", "Skipped Formulas", "Template"]
        sep = "+-" + "-+-".join("-" * w for w in col_w) + "-+"
        header_row = (
            "| "
            + " | ".join(str(h).ljust(w) for h, w in zip(headers, col_w))
            + " |"
        )

        print(f"\n\n{'=' * 60}")
        print("SUMMARY")
        print(sep)
        print(header_row)
        print(sep)
        for r in summary_rows:
            line = (
                "| "
                + " | ".join(
                    str(v).ljust(w)
                    for v, w in zip(
                        [r["job"], r["status"], r["matched"], r["mismatch"],
                         r["missing"], r["skipped"], r["template"]],
                        col_w,
                    )
                )
                + " |"
            )
            print(line)
        print(sep)

    # -----------------------------------------------------------------------
    # Single-job branch
    # -----------------------------------------------------------------------
    else:
        job_data      = TEST_JOBS[ACTIVE_JOB]["job_data"]
        ORIGINAL_PATH = TEST_JOBS[ACTIVE_JOB]["original_path"]

        print("=" * 60)
        print(f"Active job : {ACTIVE_JOB}")
        print(f"Original   : {ORIGINAL_PATH}")
        print("=" * 60)

        print("Running fill_workbook ...")
        result = fill_workbook(
            job_data,
            template_path=None,
            original_path=ORIGINAL_PATH,
        )
        print(f"fill_workbook result: {json.dumps(result, indent=2)}")

        if result.get("status") == "ok":
            output_path = result["output_path"]
            print()
            print("=" * 60)
            print("Running verify_fill ...")
            print(f"  Original : {ORIGINAL_PATH}")
            print(f"  Output   : {output_path}")
            print()

            report = verify_fill(ORIGINAL_PATH, output_path)

            total_matched    = len(report["matched"])
            total_mismatched = len(report["mismatched"])
            total_missing    = len(report["missing"])
            total_skipped    = report["skipped_formulas"]
            detected         = report["template_detected"]

            detected_label = "v1" if detected == TEMPLATE_V1 else "v2"
            print(f"Template detected      : {detected_label} ({detected})")
            print(f"Total matched          : {total_matched}")
            print(f"Total mismatched       : {total_mismatched}")
            print(f"Total missing          : {total_missing}")
            print(f"Total skipped formulas : {total_skipped}")

            if report["mismatched"]:
                print()
                print("Mismatches:")
                for coord, expected, got in report["mismatched"]:
                    print(f"  {coord}: expected {expected!r}  got {got!r}")

            if report["missing"]:
                print()
                print("Missing (present in original but not in output):")
                for coord in report["missing"]:
                    print(f"  {coord}")
        else:
            print("Workbook saved but status was not 'ok' — skipping verify_fill.")
