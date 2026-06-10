# fastapi_app.py
# ─────────────────────────────────────────────────────────────────────────────
# GFS Drawing Processor — Full Pipeline Backend
#
# This file is the brain of the entire system.
# When someone uploads a GFS shop drawing (PDF or ZIP), this file:
#   1. Validates the file type
#   2. Saves it to a temporary folder
#   3. Converts every page to an image
#   4. Sends each image to Azure AI to classify what type of page it is
#   5. Extracts structured data from the important pages
#   6. Returns everything as clean JSON
#   7. Lets the user download a formatted Excel report
#
# Endpoints:
#   GET  /health          — confirms server is alive
#   POST /upload          — main entry point, runs full pipeline, returns JSON
#   POST /select          — continues pipeline when user picks from multiple PDFs
#   POST /extract         — runs extraction on pre-classified pages
#   GET  /report/{job_id} — downloads Excel report for a completed job
# ─────────────────────────────────────────────────────────────────────────────

import os
import uuid
import shutil
import threading
import tempfile
import base64
import json
from fastapi.staticfiles import StaticFiles


from openai import OpenAI
from fastapi import FastAPI, UploadFile, File, BackgroundTasks
from fastapi.responses import JSONResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from dotenv import load_dotenv

# load_dotenv() reads the .env file and loads all variables into the environment
# This must happen before any os.getenv() calls below
# Without this, AZURE_OPENAI_KEY and other variables would be empty
load_dotenv()

# ── AZURE OPENAI CONFIGURATION ────────────────────────────────────────────────
# These four variables are the credentials and settings needed to talk to Azure AI
# They are read from the .env file — never hardcoded in the source code
# If any of these are wrong or missing, API calls will fail with auth errors
AZURE_KEY        = os.getenv("AZURE_OPENAI_KEY", "")           # the API password
AZURE_ENDPOINT   = os.getenv("AZURE_OPENAI_ENDPOINT", "").rstrip("/") + "/"  # the Azure URL
AZURE_DEPLOYMENT = os.getenv("AZURE_OPENAI_CHAT_DEPLOYMENT", "gpt-5-2")      # which AI model to use
AZURE_VERSION    = os.getenv("AZURE_OPENAI_API_VERSION", "2025-08-07")        # API version

# ── OUR OWN PIPELINE FILES ────────────────────────────────────────────────────
# These are the other Python files in this project that each handle one job:
#
# error_handler         → defines standard error responses with correct HTTP codes
#                         so all errors look consistent across the API
#
# page_classifier_ayush → takes a page image and asks Azure AI what type it is
#                         returns COVER, PLAN_VIEW, MGDS_CATALOG, or OTHER
#
# pdf_to_image          → opens a PDF file and converts each page to a PNG image
#                         uses PyMuPDF (fitz) library under the hood
#
# folder_processor      → when a ZIP is uploaded, searches inside for the GFS drawing
#                         handles cases where the ZIP has subfolders or multiple PDFs
from error_handler import (
    raise_unsupported_file_type,   # called when someone uploads a .txt or .doc etc
    raise_pdf_corrupt,             # called when the PDF can't be opened or read
    raise_no_gfs_drawing,          # called when a ZIP has no PDF inside
    ambiguous_drawings_response,   # called when a ZIP has too many PDFs to auto-pick
    cad_file_detected_response     # called when someone uploads a .dwg or .dxf CAD file
)
from page_classifier_ayush import classify_all_pages
import pdf_to_image
import folder_processor


# ── FASTAPI APP INSTANCE ──────────────────────────────────────────────────────
# This creates the web application object
# Everything — endpoints, middleware, settings — attaches to this `app` object
# The title and description show up automatically at localhost:8000/docs
# so any team member can open that URL and test the API without writing any code
app = FastAPI(
    title="GFS Document Processing API",
    description="Full pipeline — PDF upload → page classification → parameter extraction for GFS estimation.",
    version="2.0.0"
)

# ── CORS MIDDLEWARE ───────────────────────────────────────────────────────────
# CORS (Cross-Origin Resource Sharing) controls which websites can call this API
# Without this, a browser running on port 3000 would be blocked from calling
# this API on port 8000 — browsers enforce this security rule by default
# allow_origins=["*"] means any website can call this API
# This is fine for development and internal tools — lock it down in production
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"]
)

# ── CONSTANTS ─────────────────────────────────────────────────────────────────
# These values are defined once at the top so they're easy to find and change
# If you need to add a new file type or change the cleanup time, change it here
# and it automatically applies everywhere in the code

ALLOWED_EXTENSIONS = {".pdf", ".zip"}  # only these file types will be processed
CAD_EXTENSIONS     = {".dwg", ".dxf"}  # CAD files get a friendly message instead of an error
CLEANUP_DELAY_SEC  = 3600              # temp files are deleted 1 hour after processing
TEMP_BASE          = tempfile.gettempdir()
# tempfile.gettempdir() finds the right temp folder for the current operating system:
# Windows → C:\Users\ayush\AppData\Local\Temp
# Mac/Linux → /tmp

# ── IN-MEMORY JOB STORE ───────────────────────────────────────────────────────
# This is a simple Python dictionary that acts as a short-term memory for the server
# When /upload finishes, it saves the full pipeline result here using job_id as the key
# When /report is called, it looks up the result here using the same job_id
# This means the user doesn't need to re-upload the PDF just to get the Excel report
# Important: this is cleared every time the server restarts
#            for a production system, you'd use a database instead
JOB_RESULTS = {}


# ── ENDPOINT 1: HEALTH CHECK ──────────────────────────────────────────────────
# The simplest possible endpoint — just returns {"status": "ok"}
# Used by monitoring tools, load balancers, and other team members
# to confirm the server is running before sending real requests
# If this doesn't return 200, something is seriously wrong with the server
@app.get("/health", summary="Health check")
def health():
    return {"status": "ok"}


# ── ENDPOINT 2: UPLOAD ────────────────────────────────────────────────────────
# This is the main entry point for the entire system
# Everything starts here when someone submits a drawing
#
# It handles 5 different scenarios automatically:
#   1. Wrong file type (.doc, .txt etc) → return 415 error immediately
#   2. CAD file (.dwg/.dxf)            → return helpful message asking for PDF
#   3. ZIP with multiple PDFs          → return candidates list, user picks via /select
#   4. ZIP with one PDF                → run full pipeline on that PDF
#   5. Direct PDF upload               → run full pipeline on the PDF
#
# The BackgroundTasks parameter lets us schedule cleanup to happen
# AFTER the response is already sent — so the user gets results instantly
# and the cleanup runs silently in the background
@app.post("/upload", summary="Upload a PDF or ZIP — runs full pipeline, returns JSON")
async def upload(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(..., description="PDF or ZIP containing GFS shop drawings")
    # UploadFile is FastAPI's built-in file upload handler
    # File(...) — the three dots mean this field is required, request fails without it
):
    # ── Validate that the upload actually has a filename ──────────────────────
    # This is an edge case but can happen with some HTTP clients
    # Better to catch it early with a clear message than crash later
    if not file.filename:
        return JSONResponse(
            status_code=422,
            content={"error_code": "INVALID_FILENAME", "message": "Uploaded file must have a filename.", "details": {}}
        )

    # ── Extract and normalise the file extension ──────────────────────────────
    # os.path.splitext("drawing.PDF") returns ("drawing", ".PDF")
    # We only need the extension part, so we discard the name with _
    # .lower() makes sure ".PDF" and ".pdf" are treated the same
    _, ext = os.path.splitext(file.filename)
    ext = ext.lower()

    # ── Handle CAD files with a helpful message ───────────────────────────────
    # CAD files are a very common mistake — the client sends .dwg instead of PDF
    # Instead of a confusing generic error, we return a specific helpful message
    # This check must come BEFORE the general allowed check below
    # so .dwg gets the specific CAD message, not the generic unsupported error
    if ext in CAD_EXTENSIONS:
        return JSONResponse(
            status_code=200,
            content=cad_file_detected_response(details={"filename": file.filename})
        )

    # ── Reject anything that's not PDF or ZIP ─────────────────────────────────
    # HTTP 415 = Unsupported Media Type — this is the correct code for wrong format
    # raise_unsupported_file_type raises an HTTPException which stops execution here
    if ext not in ALLOWED_EXTENSIONS:
        raise_unsupported_file_type(
            details={"filename": file.filename, "received_extension": ext}
        )

    # ── Create a unique job ID and temp folder for this upload ────────────────
    # uuid4() generates a random 128-bit ID like "b7a6a091-914e-4985-aba1-2ddca2c0e1e1"
    # The collision probability is essentially zero — two jobs will never get the same ID
    # Each job gets its own folder so parallel uploads never interfere with each other
    job_id      = str(uuid.uuid4())
    temp_folder = os.path.join(TEMP_BASE, f"gfs_{job_id}")
    os.makedirs(temp_folder, exist_ok=True)
    # exist_ok=True prevents a crash if the folder somehow already exists

    saved_file_path = os.path.join(temp_folder, file.filename)
    # os.path.join is used instead of string concatenation
    # it handles path separators correctly on both Windows (\) and Linux (/)

    # ── Save the uploaded file to disk ────────────────────────────────────────
    # shutil.copyfileobj copies in small chunks instead of loading the whole file
    # this is important for large PDFs (100MB+) — we don't want to run out of memory
    # "wb" = write binary mode — PDFs and ZIPs are binary, not text
    try:
        with open(saved_file_path, "wb") as f:
            shutil.copyfileobj(file.file, f)
    except Exception as e:
        # Saving failed — delete the empty folder we just created
        # Never leave orphaned folders on disk when something goes wrong
        _cleanup_now(temp_folder)
        raise_pdf_corrupt(details={"filename": file.filename, "error": str(e)})

    # ── Route based on whether it's a ZIP or direct PDF ──────────────────────
    if ext == ".zip":
        # ZIP files need to be extracted first
        # shutil.unpack_archive is like right-clicking and selecting "Extract Here"
        # It handles .zip, .tar.gz, .tar.bz2 and other archive formats automatically
        try:
            shutil.unpack_archive(saved_file_path, temp_folder)
        except Exception as e:
            _cleanup_now(temp_folder)
            raise_pdf_corrupt(details={"filename": file.filename, "error": str(e)})

        # Sanitize filenames after extraction
        # Some ZIPs contain files with colons in the name (e.g. "drawing 9:19:25.pdf")
        # Colons are illegal in Windows file paths and cause errors
        # _sanitize_filenames replaces illegal characters with dashes
        _sanitize_filenames(temp_folder)

        # Ask folder_processor to search the extracted folder for the GFS drawing
        # It walks through all subfolders looking for PDF files
        # Returns a dict: {"ambiguous": bool, "candidates": [...], "pdf_path": "..."}
        result = folder_processor.process_zip(saved_file_path)
        print("=" * 50)
        print("ZIP RESULT:", result)
        print("=" * 50)

        # If multiple PDFs were found, we can't automatically pick the right one
        # Return all candidates so the user can choose via POST /select
        # We still schedule cleanup in case the user never comes back
        if result.get("ambiguous"):
            background_tasks.add_task(_cleanup, temp_folder)
            return JSONResponse(
                status_code=200,
                content={
                    "status":     "need_selection",
                    "job_id":     job_id,      # user needs this to call /select later
                    "candidates": result.get("candidates", [])
                }
            )

        # If no PDF was found at all inside the ZIP, raise an error
        if not result.get("selected_pdf"):
            _cleanup_now(temp_folder)
            raise_no_gfs_drawing(details={"folder": temp_folder})

        pdf_path = result["selected_pdf"]

    else:
        # Direct PDF upload — the file we just saved IS the drawing
        # No need to search or extract anything
        pdf_path = saved_file_path

    # ── Run the full pipeline ─────────────────────────────────────────────────
    # Both ZIP and direct PDF paths converge here
    # pdf_path always points to the correct drawing PDF regardless of how it arrived
    # _run_full_pipeline handles: convert to images → classify → extract → return
    response = _run_full_pipeline(
        pdf_path=pdf_path,
        job_id=job_id,
        temp_folder=temp_folder
    )

    # Schedule temp folder cleanup after 1 hour
    # background_tasks runs this AFTER the response is already sent to the user
    # The user gets their result immediately — cleanup happens silently after
    background_tasks.add_task(_cleanup, temp_folder)

    # Add the report URL to the response so the user knows how to download Excel
    # They just need to call GET /report/{job_id} in their browser or Swagger
    response["report_url"] = f"/report/{job_id}"

    return JSONResponse(status_code=200, content=response)


# ── ENDPOINT 3: SELECT ────────────────────────────────────────────────────────
# This endpoint only exists for one specific situation:
# When /upload found multiple PDFs inside a ZIP and returned "need_selection"
#
# The user sees the candidates list, picks the right one, and calls this endpoint
# with the job_id and the path of their chosen PDF
# We then run the exact same pipeline as /upload would have
#
# Example request body:
# {
#   "job_id": "b7a6a091-...",
#   "selected_pdf": "C:\Temp\gfs_b7a6a091\Approved Documents\shop_drawing.pdf"
# }
@app.post("/select", summary="Select a PDF from multiple candidates")
async def select(body: dict, background_tasks: BackgroundTasks):
    # body is the JSON request body, automatically parsed into a Python dict by FastAPI
    job_id       = body.get("job_id")
    selected_pdf = body.get("selected_pdf")

    # Both fields are required — return a clear error if either is missing
    if not job_id or not selected_pdf:
        return JSONResponse(
            status_code=422,
            content={"error_code": "INVALID_REQUEST", "message": "job_id and selected_pdf are required.", "details": {}}
        )

    # Check the selected file still exists on disk
    # It might be gone if the job expired and the 1-hour cleanup already ran
    if not os.path.exists(selected_pdf):
        return JSONResponse(
            status_code=422,
            content={"error_code": "FILE_NOT_FOUND", "message": "Selected PDF not found. Job may have expired.", "details": {"selected_pdf": selected_pdf}}
        )

    # Reconstruct the temp folder path using the job_id
    # We always name temp folders "gfs_{job_id}" when creating them in /upload
    # So we can rebuild the path here without needing to store it anywhere
    temp_folder = os.path.join(TEMP_BASE, f"gfs_{job_id}")

    # Run the same pipeline as /upload — shared helper keeps the code DRY
    # (DRY = Don't Repeat Yourself — same logic defined once, used in multiple places)
    response = _run_full_pipeline(
        pdf_path=selected_pdf,
        job_id=job_id,
        temp_folder=temp_folder
    )

    background_tasks.add_task(_cleanup, temp_folder)

    # Same as /upload — tell the user where to download the Excel report
    response["report_url"] = f"/report/{job_id}"

    return JSONResponse(status_code=200, content=response)


# ── ENDPOINT 4: EXTRACT ───────────────────────────────────────────────────────
# This endpoint skips the upload and classification steps
# It accepts a list of already-classified pages and runs extraction only
#
# Useful when:
# - You already ran /upload and want to re-extract without re-uploading
# - Classification was done separately upstream and you just need extraction
# - You want to test extraction on specific pages without a full pipeline run
#
# Example request body:
# {
#   "job_id": "b7a6a091-...",
#   "pages": [
#     {"page_number": 1, "page_type": "COVER", "confidence": "HIGH", "image_path": "C:\Temp\..."}
#   ]
# }
@app.post("/extract", summary="Extract parameters from pre-classified pages")
async def extract(body: dict):
    job_id = body.get("job_id")
    pages  = body.get("pages", [])
    # pages defaults to an empty list (not None) so len(pages) never crashes

    # job_id is required for tracking — fail clearly if missing
    if not job_id:
        return JSONResponse(
            status_code=422,
            content={"error_code": "INVALID_REQUEST", "message": "job_id is required.", "details": {}}
        )

    # Can't extract from an empty list — fail clearly
    if not pages:
        return JSONResponse(
            status_code=422,
            content={"error_code": "NO_PAGES", "message": "pages list is empty.", "details": {}}
        )

    extraction_result = _extract_from_pages(pages)

    return JSONResponse(
        status_code=200,
        content={
            "status":     "extracted",
            "job_id":     job_id,
            "page_count": len(pages),
            "extraction": extraction_result
        }
    )


# ── ENDPOINT 5: REPORT ────────────────────────────────────────────────────────
# Downloads a formatted Excel report for a job that was already run
#
# Flow:
#   Step 1: POST /upload  → get back job_id in the response
#   Step 2: GET /report/{job_id} → Excel file downloads in the browser
#
# The result from /upload is stored in JOB_RESULTS (the in-memory dict above)
# so this endpoint can generate the Excel without re-running the pipeline
#
# The Excel file has 4 sheets:
#   Sheet 1 — Overview:        job info, page classification table, any errors
#   Sheet 2 — Cover Page:      what a cover page is + all extracted fields in tables
#   Sheet 3 — Plan View Pages: what plan views are + exposed frame, panels, flags
#   Sheet 4 — Skipped & Errors: pages that were skipped or failed, with reasons
@app.get(
    "/report/{job_id}",
    summary="Download Excel report for a completed job",
    description="""
    Pass the job_id returned by /upload or /select.
    Returns a downloadable Excel file with 4 sheets:
    - Overview: job info and page classification summary
    - Cover Page: all extracted cover page fields in tables
    - Plan View Pages: exposed frame, panel info, flags per plan page
    - Skipped & Errors: skipped pages and any extraction errors
    """
)
async def report(job_id: str):
    # Check that this job_id exists in our in-memory store
    # It won't exist if the server was restarted or if the job_id is wrong
    if job_id not in JOB_RESULTS:
        return JSONResponse(
            status_code=404,
            content={
                "error": "Job not found.",
                "detail": "Either the job_id is invalid or the server was restarted. Please run /upload again."
            }
        )

    data        = JOB_RESULTS[job_id]
    report_path = os.path.join(TEMP_BASE, f"gfs_report_{job_id}.xlsx")

    # Build the Excel file from the stored pipeline result
    _generate_excel_report(data, report_path)

    # FileResponse tells FastAPI to send the file as a download
    # The browser will show a "Save As" dialog or auto-download it
    # media_type tells the browser this is an Excel file, not plain text
    return FileResponse(
        path=report_path,
        filename=f"gfs_report_{job_id}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
app.mount("/", StaticFiles(directory="static", html=True), name="static")


# ── EXCEL REPORT GENERATOR ────────────────────────────────────────────────────
# Takes the pipeline output dict and builds a structured Excel workbook
# Uses openpyxl library to create the file with formatting, colors, and merged cells
#
# The report is designed to be readable by non-technical users (estimators, managers)
# Each sheet starts with a plain-English description of what that page type means
# then shows the extracted data in color-coded tables
#
# Color coding:
#   Dark blue headers    → section titles
#   Mid blue rows        → table column headers
#   Light green rows     → glass specification data
#   Light orange rows    → frame specification data
#   Light blue rows      → project header and plan panel data
#   Orange/red text      → missing values or errors that need attention
#   Grey alternating     → alternating row colors for readability
def _generate_excel_report(data: dict, output_path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # ── Color palette — all hex codes without the # prefix ───────────────────
    DARK_BLUE    = "1F3864"   # dark navy — used for main title bars
    MID_BLUE     = "2E75B6"   # medium blue — used for table headers
    LIGHT_BLUE   = "D6E4F0"   # pale blue — used for project header fields
    ORANGE       = "C55A11"   # burnt orange — used for error headers
    LIGHT_ORANGE = "FCE4D6"   # pale orange — used for frame fields and error rows
    LIGHT_GREEN  = "E2EFDA"   # pale green — used for glass specification fields
    GREY         = "F2F2F2"   # light grey — alternating row color
    WHITE        = "FFFFFF"   # white — alternating row color
    DARK_GREY    = "595959"   # dark grey — used for subtitle text

    # ── Helper: thin border around every cell ─────────────────────────────────
    # Called as border() (not border) because it creates a new object each time
    # openpyxl requires separate Border objects — sharing one causes styling bugs
    def border():
        s = Side(style="thin", color="BFBFBF")
        return Border(left=s, right=s, top=s, bottom=s)

    # ── Helper: header cell — bold white text on colored background ───────────
    # Used for table column headers (e.g. "Field", "Extracted Value", "Page #")
    def hcell(ws, row, col, val, bg=DARK_BLUE, fg=WHITE, size=10, bold=True):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(name="Arial", bold=bold, color=fg, size=size)
        c.fill      = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = border()
        return c

    # ── Helper: data cell — regular black text on white or grey background ────
    # Used for all extracted data values in the tables
    def dcell(ws, row, col, val, bg=WHITE, bold=False, wrap=True, align="left", color="000000"):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(name="Arial", bold=bold, color=color, size=10)
        c.fill      = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        c.border    = border()
        return c

    # ── Helper: section title bar — spans multiple columns ────────────────────
    # Used for section labels like "Project Header", "Glass Specification" etc
    # The span parameter controls how many columns it merges across
    def section(ws, row, col, title, span):
        c = ws.cell(row=row, column=col, value=title)
        c.font      = Font(name="Arial", bold=True, color=WHITE, size=11)
        c.fill      = PatternFill("solid", start_color=MID_BLUE)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = border()
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=span)
        ws.row_dimensions[row].height = 22

    # ── Helper: description block — italic text explaining a page type ────────
    # Shown at the top of Cover Page and Plan View sheets
    # Explains to the reader (estimator/manager) why this page type matters
    def desc_block(ws, row, col, span, text):
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=span)
        c = ws.cell(row=row, column=col, value=text)
        c.font      = Font(name="Arial", size=10, italic=True)
        c.fill      = PatternFill("solid", start_color=LIGHT_BLUE)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border    = border()
        ws.row_dimensions[row].height = 70

    # ── Helper: page title bar — large white text on dark blue ────────────────
    # The big heading at the top of each sheet
    def title_row(ws, row, col, span, text):
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=span)
        c = ws.cell(row=row, column=col, value=text)
        c.font      = Font(name="Arial", bold=True, size=16, color=WHITE)
        c.fill      = PatternFill("solid", start_color=DARK_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 36

    # ── Helper: set column widths for a sheet ────────────────────────────────
    # Takes a list of (column_letter, width) pairs
    # e.g. [("A", 5), ("B", 30), ("C", 40)]
    def set_cols(ws, widths):
        for col, w in widths:
            ws.column_dimensions[col].width = w

    # ── Helper: a label + value row spanning multiple columns ─────────────────
    # Used for every extracted field — label on left, value merged across remaining cols
    # Missing values show in red so the estimator can spot gaps immediately
    def field_row(ws, row, col, label, value, label_bg=LIGHT_BLUE, span_to=5):
        dcell(ws, row, col, label, bg=label_bg, bold=True, align="right")
        color = "C00000" if value in (None, "", "Not found") else "000000"
        val   = str(value) if value not in (None, "") else "Not found"
        c = ws.cell(row=row, column=col+1, value=val)
        c.font      = Font(name="Arial", size=10, color=color)
        c.fill      = PatternFill("solid", start_color=WHITE)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border    = border()
        ws.merge_cells(start_row=row, start_column=col+1, end_row=row, end_column=span_to)
        ws.row_dimensions[row].height = 20

    # Create a new empty workbook and remove the default empty sheet
    wb = Workbook()
    wb.remove(wb.active)

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 1 — OVERVIEW
    # Purpose: give the reader a quick summary of the entire job at a glance
    # Shows: job ID, status, page counts, and a classification table for every page
    # Also shows any extraction errors at the bottom so nothing is hidden
    # ─────────────────────────────────────────────────────────────────────────
    ws1 = wb.create_sheet("Overview")
    ws1.sheet_view.showGridLines = False  # cleaner look without the grey grid
    set_cols(ws1, [("A",5),("B",28),("C",40),("D",18),("E",18),("F",18)])

    # Main title
    title_row(ws1, 1, 2, 6, "GFS Drawing Extraction Report")

    # Subtitle
    ws1.merge_cells("B2:F2")
    sub = ws1["B2"]
    sub.value = "Glass Flooring Systems Inc — Automated Pipeline Output"
    sub.font  = Font(name="Arial", size=11, color=DARK_GREY, italic=True)
    sub.fill  = PatternFill("solid", start_color=LIGHT_BLUE)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 22

    # Job information section — shows the technical details of this specific run
    row = 4
    section(ws1, row, 2, "  Job Information", 6); row += 1
    summary = data.get("extraction", {}).get("summary", {})
    for label, value in [
        ("Job ID",        data.get("job_id", "N/A")),
        ("Status",        data.get("status", "N/A").upper()),
        ("Total Pages",   summary.get("total_pages", 0)),
        ("Cover Pages",   summary.get("cover_pages", 0)),
        ("Plan Pages",    summary.get("plan_view_pages", 0)),
        ("Skipped Pages", summary.get("skipped_pages", 0)),
        ("Failed Pages",  summary.get("failed_pages", 0)),
    ]:
        dcell(ws1, row, 2, label, bg=LIGHT_BLUE, bold=True, align="right")
        dcell(ws1, row, 3, str(value))
        ws1.row_dimensions[row].height = 20; row += 1

    # Page classification table — one row per page showing what the AI classified it as
    # Color coded by type so the reader can instantly see the mix of pages
    row += 1
    section(ws1, row, 2, "  Page Classification Summary", 6); row += 1
    for col, h in [(2,"Page #"),(3,"Page Type"),(4,"Confidence"),(5,"Description")]:
        hcell(ws1, row, col, h, bg=MID_BLUE)
    ws1.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    ws1.row_dimensions[row].height = 20; row += 1

    # Each page type gets its own background color for easy scanning
    type_colors = {
        "COVER":        LIGHT_GREEN,    # green = most important, extraction happens here
        "PLAN_VIEW":    LIGHT_BLUE,     # blue = also important, dimensions come from here
        "MGDS_CATALOG": LIGHT_ORANGE,   # orange = catalog page, handled separately
        "OTHER":        GREY            # grey = skipped, not useful for estimation
    }
    # Human-readable description of each page type shown in the table
    type_desc = {
        "COVER":        "Main project page — glass specs, frame type, contractor info",
        "PLAN_VIEW":    "Top-down layout drawing — panel dimensions and exposed frame",
        "MGDS_CATALOG": "Modular Glass Deck System catalog reference page",
        "OTHER":        "Section cut, detail drawing, warranty, label — skipped",
    }
    for page in data.get("pages", []):
        pt = page.get("page_type", "OTHER")
        bg = type_colors.get(pt, WHITE)
        dcell(ws1, row, 2, page.get("page_number"), bg=bg, align="center")
        dcell(ws1, row, 3, pt,                      bg=bg, bold=True)
        dcell(ws1, row, 4, page.get("confidence"),  bg=bg, align="center")
        c = ws1.cell(row=row, column=5, value=type_desc.get(pt, ""))
        c.font      = Font(name="Arial", size=10)
        c.fill      = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border    = border()
        ws1.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        ws1.row_dimensions[row].height = 20; row += 1

    # Extraction errors section — only shown if something failed
    # Errors are shown in orange/red so they're impossible to miss
    errors = data.get("extraction", {}).get("errors", [])
    if errors:
        row += 1
        section(ws1, row, 2, "  Extraction Errors", 6); row += 1
        for col, h in [(2,"Page #"),(3,"Page Type"),(4,"Error")]:
            hcell(ws1, row, col, h, bg=ORANGE)
        ws1.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        ws1.row_dimensions[row].height = 20; row += 1
        for err in errors:
            dcell(ws1, row, 2, err.get("page_number"), bg=LIGHT_ORANGE, align="center")
            dcell(ws1, row, 3, err.get("page_type"),   bg=LIGHT_ORANGE)
            c = ws1.cell(row=row, column=4, value=err.get("error",""))
            c.font      = Font(name="Arial", size=10, color="C00000")
            c.fill      = PatternFill("solid", start_color=LIGHT_ORANGE)
            c.alignment = Alignment(wrap_text=True, vertical="center")
            c.border    = border()
            ws1.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
            ws1.row_dimensions[row].height = 30; row += 1

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 2 — COVER PAGE
    # Purpose: show everything extracted from the cover page in structured tables
    # The cover page is the most data-rich page — project info, glass specs, frame, units
    # Starts with a plain-English description so non-technical readers understand it
    # ─────────────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Cover Page")
    ws2.sheet_view.showGridLines = False
    set_cols(ws2, [("A",5),("B",30),("C",40),("D",20),("E",20)])

    title_row(ws2, 1, 2, 5, "Cover Page Extraction")

    # Description block — explains what a cover page is to anyone reading this report
    row = 3
    section(ws2, row, 2, "  What is a Cover Page?", 5); row += 1
    desc_block(ws2, row, 2, 5,
        "The COVER page is the most critical page in a GFS shop drawing. It contains all the "
        "project-level information needed to begin estimation: the project address, contractor "
        "details, glass makeup specification (number of layers, type, texture), frame series, "
        "finish, cap color, and the full units table with dimensions and quantities. "
        "Every field extracted here directly feeds into the cost estimation workflow."
    ); row += 2

    cover = data.get("extraction", {}).get("cover")

    # If no cover was extracted (all pages were OTHER or extraction failed)
    if not cover:
        ws2.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        c = ws2.cell(row=row, column=2, value="No cover page was extracted from this drawing.")
        c.font = Font(name="Arial", size=11, color="C00000", bold=True)
    else:
        # Unpack the nested data structure from the pipeline output
        d     = cover.get("data", {})
        ph    = d.get("project_header", {}) or {}       # project address, contractor etc
        gs    = d.get("glass_specification", {}) or {}  # glass makeup, back paint etc
        fr    = d.get("frame", {}) or {}                 # series, finish, cap color etc
        units = d.get("units", []) or []                 # unit dimensions table rows

        # Show which page number this came from and its confidence level
        dcell(ws2, row, 2, "Page Number", bg=LIGHT_BLUE, bold=True, align="right")
        dcell(ws2, row, 3, cover.get("page_number","N/A"), align="center")
        dcell(ws2, row, 4, "Confidence",  bg=LIGHT_BLUE, bold=True, align="right")
        dcell(ws2, row, 5, cover.get("confidence","N/A"), align="center")
        ws2.row_dimensions[row].height = 20; row += 2

        # ── Project Header table ──────────────────────────────────────────────
        # Shows the who/what/where of the project
        # Missing fields show in red so the estimator knows what to fill in manually
        section(ws2, row, 2, "  Project Header", 5); row += 1
        hcell(ws2, row, 2, "Field",           bg=MID_BLUE)
        hcell(ws2, row, 3, "Extracted Value", bg=MID_BLUE)
        ws2.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
        ws2.row_dimensions[row].height = 20; row += 1
        for label, value in [
            ("Project Address", ph.get("project_address")),
            ("Contractor",      ph.get("contractor")),
            ("Manufacturer",    ph.get("manufacturer")),
            ("Project Name",    ph.get("project_name")),
            ("Drawing Date",    ph.get("drawing_date")),
            ("Sheet Number",    ph.get("sheet_number")),
            ("Quote Number",    ph.get("quote_number")),
            ("Revision",        ph.get("revision")),
        ]:
            field_row(ws2, row, 2, label, value, label_bg=LIGHT_BLUE, span_to=5)
            row += 1

        # ── Glass Specification table ─────────────────────────────────────────
        # Shows the glass makeup — this is critical for material costing
        # EXPEDITED flag is highlighted in orange because it affects lead time and pricing
        row += 1
        section(ws2, row, 2, "  Glass Specification", 5); row += 1
        hcell(ws2, row, 2, "Field",           bg=MID_BLUE)
        hcell(ws2, row, 3, "Extracted Value", bg=MID_BLUE)
        ws2.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
        ws2.row_dimensions[row].height = 20; row += 1
        for label, value in [
            ("Glass Makeup",      gs.get("glass_makeup")),
            ("Glass Type",        gs.get("glass_type")),
            ("Top Layer Texture", gs.get("top_layer_texture")),
            ("Back Paint",        gs.get("back_paint")),
            ("HST",               gs.get("hst")),
            ("Expedited",         gs.get("expedited")),
        ]:
            is_expedited = label == "Expedited" and str(value).lower() == "yes"
            field_row(ws2, row, 2, label, value, label_bg=LIGHT_GREEN, span_to=5)
            if is_expedited:
                # Override the font color to orange for EXPEDITED = yes
                # This makes it visually stand out so the estimator doesn't miss it
                c = ws2.cell(row=row, column=3)
                c.font = Font(name="Arial", size=10, color="C55A11", bold=True)
            ws2.row_dimensions[row].height = 30 if label == "Glass Makeup" else 20
            row += 1

        # ── Frame Specification table ─────────────────────────────────────────
        # Shows the frame type, finish, colors — all needed for the cost estimate
        row += 1
        section(ws2, row, 2, "  Frame Specification", 5); row += 1
        hcell(ws2, row, 2, "Field",           bg=MID_BLUE)
        hcell(ws2, row, 3, "Extracted Value", bg=MID_BLUE)
        ws2.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
        ws2.row_dimensions[row].height = 20; row += 1
        for label, value in [
            ("Series / Type",   fr.get("series") or fr.get("perimeter_frame_type")),
            ("Frame Material",  fr.get("frame_material")),
            ("Finish",          fr.get("finish") or fr.get("frame_finish")),
            ("Cap Color",       fr.get("cap_color")),
            ("Frame Color",     fr.get("frame_color")),
            ("Cross Beam Type", fr.get("cross_beam_type")),
            ("Frame Assembly",  fr.get("frame_assembly")),
        ]:
            field_row(ws2, row, 2, label, value, label_bg=LIGHT_ORANGE, span_to=5)
            row += 1

        # ── Units / Dimensions table ──────────────────────────────────────────
        # Shows the dimensions table from the drawing — one row per unit
        # Width and length are kept as raw strings (e.g. '47-5/8"') not converted
        # because downstream estimation tools need the original format
        row += 1
        section(ws2, row, 2, "  Units / Dimensions Table", 5); row += 1
        if not units:
            # Some drawings don't have a units table on the cover page — that's normal
            ws2.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
            c = ws2.cell(row=row, column=2, value="No units table found on this cover page.")
            c.font      = Font(name="Arial", size=10, italic=True, color=DARK_GREY)
            c.fill      = PatternFill("solid", start_color=GREY)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = border()
            ws2.row_dimensions[row].height = 20
        else:
            # Column headers for the units table
            for col_i, h in enumerate(["Unit ID","Width","Length","Quantity","Sq Ft","Perimeter","Notes"], start=2):
                hcell(ws2, row, col_i, h, bg=MID_BLUE)
            ws2.row_dimensions[row].height = 20; row += 1
            # One row per unit with alternating grey/white for readability
            for i, unit in enumerate(units):
                bg = GREY if i % 2 == 0 else WHITE
                dcell(ws2, row, 2, unit.get("unit_id"),   bg=bg, align="center", bold=True)
                dcell(ws2, row, 3, unit.get("width"),     bg=bg, align="center")
                dcell(ws2, row, 4, unit.get("length"),    bg=bg, align="center")
                dcell(ws2, row, 5, unit.get("quantity"),  bg=bg, align="center")
                dcell(ws2, row, 6, unit.get("sqft"),      bg=bg, align="center")
                dcell(ws2, row, 7, unit.get("perimeter"), bg=bg, align="center")
                dcell(ws2, row, 8, unit.get("notes"),     bg=bg)
                ws2.row_dimensions[row].height = 20; row += 1

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 3 — PLAN VIEW PAGES
    # Purpose: show everything extracted from each plan view page
    # There can be multiple plan view pages — one per unit (Unit A, Unit B etc)
    # Each plan page gets its own section with exposed frame, panel info, and flags
    # ─────────────────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Plan View Pages")
    ws3.sheet_view.showGridLines = False
    set_cols(ws3, [("A",5),("B",30),("C",35),("D",20),("E",20)])

    title_row(ws3, 1, 2, 5, "Plan View Page Extraction")

    # Description block — explains what plan views are and why they matter
    row = 3
    section(ws3, row, 2, "  What is a Plan View Page?", 5); row += 1
    desc_block(ws3, row, 2, 5,
        "PLAN VIEW pages show a top-down architectural drawing of the glass unit layout. "
        "These pages are critical for extracting EXPOSED FRAME dimensions — the visible aluminum "
        "border around the glass — which directly determines material quantities and pricing. "
        "Each Plan View page typically represents one unit (Unit A, Unit B, etc.) and shows "
        "the panel count, layout pattern, and any custom shapes that require manual estimator review."
    ); row += 2

    plans = data.get("extraction", {}).get("plans", [])

    if not plans:
        ws3.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        c = ws3.cell(row=row, column=2, value="No Plan View pages were extracted from this drawing.")
        c.font = Font(name="Arial", size=11, color="C00000", bold=True)
    else:
        # Loop through each plan view page — each gets its own section in the sheet
        for plan in plans:
            d = plan.get("data", {})

            # Section header shows which page this is and the confidence level
            section(ws3, row, 2,
                f"  Plan View — Page {plan.get('page_number')}  |  Confidence: {plan.get('confidence')}",
                5); row += 1

            # ── Exposed Frame Dimensions table ────────────────────────────────
            # The most important data on a plan view page
            # EXPOSED FRAME = the visible aluminum border — directly determines pricing
            # If none found, show a warning flag for the estimator to check manually
            hcell(ws3, row, 2, "Exposed Frame Dimensions", bg=MID_BLUE)
            hcell(ws3, row, 3, "Value (raw)",              bg=MID_BLUE)
            hcell(ws3, row, 4, "Direction",                bg=MID_BLUE)
            ws3.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
            ws3.row_dimensions[row].height = 20; row += 1

            frame_dims = d.get("exposed_frame_dimensions", []) or []
            if not frame_dims:
                # No dimensions found — this is a common issue and needs manual review
                ws3.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
                c = ws3.cell(row=row, column=2,
                    value="⚠  No EXPOSED FRAME dimensions found — verify manually.")
                c.font      = Font(name="Arial", size=10, color="C55A11", italic=True)
                c.fill      = PatternFill("solid", start_color=LIGHT_ORANGE)
                c.alignment = Alignment(wrap_text=True, vertical="center")
                c.border    = border()
                ws3.row_dimensions[row].height = 20; row += 1
            else:
                # One row per dimension with its direction (width or length)
                for i, dim in enumerate(frame_dims):
                    bg = GREY if i % 2 == 0 else WHITE
                    dcell(ws3, row, 2, f"Dim {i+1}",          bg=LIGHT_BLUE, bold=True)
                    dcell(ws3, row, 3, dim.get("value"),       bg=bg, align="center")
                    c = ws3.cell(row=row, column=4, value=dim.get("direction","unknown"))
                    c.font      = Font(name="Arial", size=10)
                    c.fill      = PatternFill("solid", start_color=bg)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    c.border    = border()
                    ws3.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
                    ws3.row_dimensions[row].height = 20; row += 1

            # ── Panel Information table ───────────────────────────────────────
            # Shows how many panels, their layout, and whether any are custom shapes
            # Custom shapes flag is shown in red because they require manual pricing
            row += 1
            section(ws3, row, 2, "  Panel Information", 5); row += 1
            for label, value in [
                ("Panel Count",      d.get("panel_count")),
                ("Panel Layout",     d.get("panel_layout")),
                ("Panel Shapes",     ", ".join(d.get("panel_shapes",[]) or [])),
                ("Has Custom Shape", str(d.get("has_custom_shape", False))),
                ("Location Label",   d.get("location_label")),
                ("Floor Level",      d.get("floor_level")),
            ]:
                field_row(ws3, row, 2, label, value, label_bg=LIGHT_BLUE, span_to=5)
                if label == "Has Custom Shape" and str(value).lower() == "true":
                    # Override to red bold — custom shapes need manual estimator attention
                    c = ws3.cell(row=row, column=3)
                    c.font = Font(name="Arial", size=10, color="C00000", bold=True)
                row += 1

            # ── Estimator Flags ───────────────────────────────────────────────
            # Warnings raised by the AI during extraction
            # Examples: missing dimensions, ambiguous panel counts, custom shapes
            # Each flag is something the estimator should verify before pricing
            flags = d.get("flags", []) or []
            if flags:
                row += 1
                section(ws3, row, 2, "  Estimator Flags", 5); row += 1
                for i, flag in enumerate(flags):
                    bg = LIGHT_ORANGE if i % 2 == 0 else WHITE
                    ws3.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
                    c = ws3.cell(row=row, column=2, value=f"⚠  {flag}")
                    c.font      = Font(name="Arial", size=10, color="C55A11")
                    c.fill      = PatternFill("solid", start_color=bg)
                    c.alignment = Alignment(wrap_text=True, vertical="center")
                    c.border    = border()
                    ws3.row_dimensions[row].height = 25; row += 1

            # ── Drawing Notes ─────────────────────────────────────────────────
            # Any text notes visible on the plan view page
            # Examples: scale, sheet references, project name, approval notes
            notes = d.get("drawing_notes", []) or []
            if notes:
                row += 1
                section(ws3, row, 2, "  Drawing Notes", 5); row += 1
                for i, note in enumerate(notes):
                    bg = GREY if i % 2 == 0 else WHITE
                    ws3.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
                    c = ws3.cell(row=row, column=2, value=note)
                    c.font      = Font(name="Arial", size=10)
                    c.fill      = PatternFill("solid", start_color=bg)
                    c.alignment = Alignment(wrap_text=True, vertical="center")
                    c.border    = border()
                    ws3.row_dimensions[row].height = 20; row += 1

            row += 2  # add some breathing space between plan page sections

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 4 — SKIPPED & ERRORS
    # Purpose: full transparency about what was NOT processed and why
    # Skipped pages = intentionally ignored (OTHER, MGDS_CATALOG)
    # Error pages = tried to extract but something went wrong
    # This sheet ensures nothing is silently lost
    # ─────────────────────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Skipped & Errors")
    ws4.sheet_view.showGridLines = False
    set_cols(ws4, [("A",5),("B",15),("C",55),("D",20)])

    title_row(ws4, 1, 2, 4, "Skipped Pages & Errors")

    # Skipped pages section — pages the pipeline chose not to extract
    # OTHER pages are things like section cuts, warranties, FedEx labels
    # These are classified correctly but skipped because they have no estimation data
    row = 3
    section(ws4, row, 2, "  Skipped Pages (OTHER / MGDS_CATALOG)", 4); row += 1
    hcell(ws4, row, 2, "Page #",  bg=MID_BLUE)
    hcell(ws4, row, 3, "Reason",  bg=MID_BLUE)
    ws4.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    ws4.row_dimensions[row].height = 20; row += 1

    for i, s in enumerate(data.get("extraction",{}).get("skipped",[]) or []):
        bg = GREY if i % 2 == 0 else WHITE
        dcell(ws4, row, 2, s.get("page_number"), bg=bg, align="center")
        c = ws4.cell(row=row, column=3, value=s.get("reason",""))
        c.font      = Font(name="Arial", size=10)
        c.fill      = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border    = border()
        ws4.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        ws4.row_dimensions[row].height = 20; row += 1

    # Errors section — pages the pipeline tried to extract but failed
    # Common causes: Azure API returned empty, image too blurry, JSON parse failed
    # The full error message is shown so the developer can debug it
    if errors:
        row += 1
        section(ws4, row, 2, "  Extraction Errors", 4); row += 1
        for col, h in [(2,"Page #"),(3,"Page Type"),(4,"Error")]:
            hcell(ws4, row, col, h, bg=ORANGE)
        ws4.row_dimensions[row].height = 20; row += 1
        for i, err in enumerate(errors):
            bg = LIGHT_ORANGE if i % 2 == 0 else WHITE
            dcell(ws4, row, 2, err.get("page_number"), bg=bg, align="center")
            dcell(ws4, row, 3, err.get("page_type"),   bg=bg)
            c = ws4.cell(row=row, column=4, value=err.get("error",""))
            c.font      = Font(name="Arial", size=10, color="C00000")
            c.fill      = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(wrap_text=True, vertical="center")
            c.border    = border()
            ws4.row_dimensions[row].height = 30; row += 1

    # Save the completed workbook to the output path
    wb.save(output_path)


# ── SANITIZE FILENAMES ────────────────────────────────────────────────────────
# Windows does not allow certain characters in file paths:
# colon (:), asterisk (*), question mark (?), quotes, angle brackets, pipe
# Some ZIP files contain PDFs with these characters in their names
# (e.g. "approved glass-182 robinson-9:19:25.pdf" — the colon causes an error)
# This function walks through all files in a folder and replaces illegal characters
# with dashes before we try to open or process any of them
def _sanitize_filenames(folder: str):
    illegal = r'\/:*?"<>|'
    for root, dirs, files in os.walk(folder, topdown=False):
        for name in files:
            clean = name
            for ch in illegal:
                clean = clean.replace(ch, "-")
            if clean != name:
                # Only rename if something actually changed
                os.rename(os.path.join(root, name), os.path.join(root, clean))


# ── CORE PIPELINE HELPER ──────────────────────────────────────────────────────
# This is the heart of the system — called by both /upload and /select
# Putting the logic here (instead of duplicating it in both endpoints) means
# any change to the pipeline only needs to be made in one place
#
# Stage 1 — pdf_to_image.convert_pdf_to_images()
#   Opens the PDF and renders each page as a high-DPI PNG
#   200 DPI is used because GFS drawings have small fractional dimensions
#   that need to be sharp enough for the AI to read accurately
#   Returns: [{"page_number": 1, "image_path": "C:\Temp\page_1.png"}, ...]
#
# Stage 2 — classify_all_pages()
#   Sends each PNG to Azure GPT Vision and asks "what type of page is this?"
#   Returns: [{"page_number": 1, "page_type": "COVER", "confidence": "HIGH", ...}, ...]
#
# Stage 3 — _extract_from_pages()
#   Loops through the classified pages
#   COVER pages → sends to Azure with the cover extraction prompt
#   PLAN_VIEW pages → sends to Azure with the plan extraction prompt
#   OTHER/MGDS_CATALOG → skipped, added to skipped list
#
# Stage 4 — Saves result to JOB_RESULTS and returns it
#   The result is stored in memory so /report can access it without re-running
def _run_full_pipeline(pdf_path: str, job_id: str, temp_folder: str) -> dict:
    # Stage 1: Convert PDF to images
    try:
        page_images = pdf_to_image.convert_pdf_to_images(
            pdf_path=pdf_path,
            output_dir=temp_folder
        )
    except Exception as e:
        # PDF conversion failed — could be corrupt, password protected, or empty
        # Clean up temp folder before raising so no files are left behind
        _cleanup_now(temp_folder)
        raise_pdf_corrupt(details={"pdf_path": pdf_path, "error": str(e)})

    # Stage 2: Classify each page
    classified_pages  = classify_all_pages(page_images)

    # Stage 3: Extract data from relevant pages
    extraction_result = _extract_from_pages(classified_pages)

    # Stage 4: Build result dict, store in memory, return
    result = {
        "status":     "extracted",
        "job_id":     job_id,
        "pages":      classified_pages,   # full classification list — all pages
        "extraction": extraction_result   # structured data from COVER and PLAN_VIEW
    }

    # Store result so /report/{job_id} can generate Excel without re-running
    JOB_RESULTS[job_id] = result

    return result


def _normalize_plan_data(raw: dict) -> dict:
    """
    Normalize plan extraction output to guaranteed flat schema.
    Handles both flat responses and nested responses where the model
    used section keys like EXPOSED_FRAME, PANELS, ROOM_LOCATION, NOTES.
    """
    # If already flat (has top-level panel_count), return with array guarantees
    if "panel_count" in raw or "exposed_frame_dimensions" in raw:
        return {
            "exposed_frame_dimensions": raw.get("exposed_frame_dimensions") or [],
            "panel_count":              raw.get("panel_count"),
            "panel_layout":             raw.get("panel_layout"),
            "panel_shapes":             raw.get("panel_shapes") or [],
            "has_custom_shape":         raw.get("has_custom_shape", False),
            "location_label":           raw.get("location_label"),
            "floor_level":              raw.get("floor_level"),
            "drawing_notes":            raw.get("drawing_notes") or [],
            "flags":                    raw.get("flags") or [],
        }

    # Model returned nested structure — flatten it
    exposed = raw.get("EXPOSED_FRAME") or raw.get("exposed_frame") or {}
    panels  = raw.get("PANELS") or raw.get("panels") or {}
    room    = raw.get("ROOM_LOCATION") or raw.get("ROOM") or raw.get("room") or {}
    notes   = raw.get("NOTES") or raw.get("notes") or {}

    return {
        "exposed_frame_dimensions": exposed.get("exposed_frame_dimensions") or [],
        "panel_count":              panels.get("panel_count"),
        "panel_layout":             panels.get("panel_layout"),
        "panel_shapes":             panels.get("panel_shapes") or [],
        "has_custom_shape":         panels.get("has_custom_shape", False),
        "location_label":           room.get("location_label"),
        "floor_level":              room.get("floor_level"),
        "drawing_notes":            notes.get("drawing_notes") or [],
        "flags":                    notes.get("flags") or [],
    }


# ── EXTRACTION HELPER ─────────────────────────────────────────────────────────
# Loops through the classified page list and extracts data from each relevant page
#
# For every page it checks three things before trying to extract:
#   1. Is it a type we should extract from? (skip OTHER and MGDS_CATALOG)
#   2. Does the image file exist on disk? (guard against missing temp files)
#   3. Can we read the image file? (guard against permissions or corruption)
#
# If all checks pass, it calls _call_extractor with the right prompt
# All results and errors are collected and returned together
def _extract_from_pages(classified_pages: list) -> dict:
    cover_data = None  # only one cover page per drawing
    plan_data  = []    # can have multiple plan pages (one per unit)
    skipped    = []    # pages we chose not to extract from
    errors     = []    # pages we tried to extract but failed

    for page in classified_pages:
        page_type   = page.get("page_type")
        page_number = page.get("page_number")
        image_path  = page.get("image_path")
        confidence  = page.get("confidence")

        # OTHER pages (section cuts, warranties, labels etc) have no useful data
        # Skip them and record why so the Excel report can show this
        if page_type == "OTHER":
            skipped.append({"page_number": page_number, "reason": "OTHER page — not relevant for extraction"})
            continue

        # MGDS catalog pages are handled by a separate workflow — skip for now
        if page_type == "MGDS_CATALOG":
            skipped.append({"page_number": page_number, "reason": "MGDS catalog page — handled separately"})
            continue

        # Guard: the image file must exist on disk
        # This shouldn't happen in normal flow but handles edge cases safely
        if not image_path or not os.path.exists(image_path):
            errors.append({"page_number": page_number, "page_type": page_type, "error": "Image file not found on disk"})
            continue

        # Read the image file as raw bytes
        # _call_extractor needs bytes, not a file path
        try:
            with open(image_path, "rb") as f:
                image_bytes = f.read()
        except Exception as e:
            errors.append({"page_number": page_number, "page_type": page_type, "error": f"Could not read image: {str(e)}"})
            continue

        # ── COVER page: extract project details, specs, frame, units ─────────
        if page_type == "COVER":
            try:
                cover_result = _call_extractor(image_bytes, _get_cover_prompt())
                cover_data   = {"page_number": page_number, "confidence": confidence, "data": cover_result}
            except Exception as e:
                errors.append({"page_number": page_number, "page_type": "COVER", "error": f"Cover extraction failed: {str(e)}"})

        # ── PLAN_VIEW page: extract exposed frame, panels, layout ────────────
        # Multiple plan pages are possible — one per unit (Unit A, Unit B etc)
        # Each result is appended to the list so none are overwritten
        elif page_type == "PLAN_VIEW":
            try:
                plan_result = _call_extractor(image_bytes, _get_plan_prompt())
                plan_result = _normalize_plan_data(plan_result)  # ← guarantee flat schema
                plan_data.append({"page_number": page_number, "confidence": confidence, "data": plan_result})
            except Exception as e:
                errors.append({"page_number": page_number, "page_type": "PLAN_VIEW", "error": f"Plan extraction failed: {str(e)}"})

    return {
        "cover":   cover_data,   # all data from the cover page
        "plans":   plan_data,    # list of data from each plan view page
        "summary": {
            "total_pages":     len(classified_pages),
            "cover_pages":     1 if cover_data else 0,
            "plan_view_pages": len(plan_data),
            "skipped_pages":   len(skipped),
            "failed_pages":    len(errors)
        },
        "skipped": skipped,      # list of skipped pages with reasons
        "errors":  errors        # list of failed pages with error messages
    }


# ── AZURE OPENAI VISION CALLER ────────────────────────────────────────────────
# This is the function that actually talks to Azure AI for extraction
# Used by both COVER and PLAN_VIEW extraction — same function, different prompt
#
# How it works:
#   1. Converts the image bytes to a base64 data URL
#      (APIs can't receive raw binary, base64 encodes it as safe text)
#   2. Creates an OpenAI client pointing to the Azure endpoint
#   3. Sends the prompt + image to GPT-5 Vision
#   4. Extracts the text response and parses it as JSON
#   5. Returns the parsed dict
#
# Uses the same OpenAI SDK approach as vision_client.py which is confirmed working
# Text prompt goes first, then image — this order improves accuracy
def _call_extractor(image_bytes: bytes, prompt: str) -> dict:
    # Convert image bytes to base64 data URL
    # Format: "data:image/png;base64,<base64_encoded_bytes>"
    b64      = base64.standard_b64encode(image_bytes).decode("utf-8")
    data_url = f"data:image/png;base64,{b64}"

    # Create OpenAI client with Azure credentials from .env
    client = OpenAI(api_key=AZURE_KEY, base_url=AZURE_ENDPOINT)

    # Send request to GPT-5 Vision
    # max_completion_tokens=4000 handles large cover pages with dense units tables
    # If set too low, the response gets cut off mid-JSON and parsing fails
    completion = client.chat.completions.create(
        model=AZURE_DEPLOYMENT,
        max_completion_tokens=4000,
        messages=[{
            "role": "user",
            "content": [
                {"type": "text",      "text": prompt},        # instructions first
                {"type": "image_url", "image_url": {"url": data_url}},  # then the image
            ]
        }]
    )

    # Extract the text content from the response object
    raw = completion.choices[0].message.content

    # Debug log — shows first 300 chars so we can see if the model is responding
    # Remove this line once the system is fully stable in production
    print(f"[DEBUG] Extractor response: {repr(raw[:300] if raw else 'EMPTY')}")

    # Guard against empty response
    # Can happen if the image is too blurry or the token limit was hit
    if not raw or not raw.strip():
        raise ValueError("Model returned empty response — check image quality or token limit")

    raw = raw.strip()

    # Strip markdown code fences if the model wrapped its JSON in ```json ... ```
    # GPT-5 sometimes adds these even when the prompt says not to
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]

    # Parse the JSON string into a Python dict
    # If this fails, the exception bubbles up to _extract_from_pages
    # which logs it as an error in the errors list
    return json.loads(raw.strip())


# ── EXTRACTION PROMPTS ────────────────────────────────────────────────────────
# Each prompt is a detailed set of instructions telling the AI exactly what to extract
# Kept as separate functions so they can be updated independently without touching
# any of the pipeline logic above
#
# Key principles in these prompts:
# - Very specific field names so the JSON structure is predictable
# - "null" for missing fields (not empty string or "N/A") for consistent parsing
# - "JSON only" at the end prevents the model from adding explanatory text
# - "Do NOT convert fractions" preserves original dimension strings like '47-5/8"'

def _get_cover_prompt() -> str:
    # Prompt for COVER pages
    # Extracts: project header, glass specification, frame details, units table
    return """You are extracting structured data from a GFS glass manufacturer shop drawing COVER page.

Extract ALL of the following fields carefully:

PROJECT HEADER:
- project_address: full street address
- contractor: contractor/client name
- drawing_date: date on drawing
- quote_number: quote or order number if visible
- revision: revision letter (e.g. Rev B)

GLASS SPECIFICATION:
- glass_makeup: full glass spec string exactly as written (e.g. "5-ply 10mm LIT NanoDot")
- glass_type: e.g. LIT, IGU, monolithic
- back_paint: yes/no and color if specified
- hst: yes/no (Heat Soak Test)
- expedited: yes/no (look for EXPEDITED text anywhere on page)

FRAME:
- series: e.g. "Series 1000", "Series 2000", "Series 3000 CityScape"
- frame_material: aluminum / steel / other
- finish: anodized / powder coat / other and color

UNITS (extract each unit row from the dimensions table):
- units: array of objects, each with:
  - unit_id: label (e.g. "Unit A", "1", etc.)
  - width: raw string exactly as written (e.g. '47-5/8"')
  - length: raw string exactly as written
  - quantity: integer
  - sqft: if shown
  - perimeter: if shown
  - notes: any special notes for this unit

Respond ONLY with valid JSON. Use null for any field not found.
Do NOT convert fractional dimensions — preserve raw strings exactly.
Do NOT include any explanation or markdown — JSON only."""


def _get_plan_prompt() -> str:
    # Prompt for PLAN_VIEW pages
    # Extracts: exposed frame dimensions, panel layout, location, estimator flags

    return """You are extracting structural data from a GFS glass manufacturer PLAN VIEW drawing.
This is a top-down architectural drawing showing glass panel layout.

STEP 1 — Read the page title carefully.
If the title contains "PLAN - GLASS UNIT", "GLASS UNIT", or "PLAN VIEW - GLASS UNIT",
this page shows glass unit dimensions. Even if dimensions are NOT labelled "EXPOSED FRAME",
extract all visible numeric dimensions as glass unit dimensions.

STEP 2 — Extract EXPOSED FRAME dimensions.
Look for dimensions explicitly labelled "EXPOSED FRAME" or "EXP. FRAME".
If found, extract them. If NOT found but visible dimensions exist and page title contains
"GLASS UNIT", extract those visible dimensions treating them as the glass unit size.

STEP 3 — Determine direction.
The SMALLER dimension is width. The LARGER dimension is length.

Return ONLY this exact flat JSON structure — no nested sections, no extra keys:

{
  "exposed_frame_dimensions": [
    {"value": "32\"", "direction": "width"},
    {"value": "96\"", "direction": "length"}
  ],
  "panel_count": 1,
  "panel_layout": "single panel",
  "panel_shapes": ["rectangle"],
  "has_custom_shape": false,
  "location_label": null,
  "floor_level": null,
  "drawing_notes": [],
  "flags": []
}

CRITICAL RULES:
- ALL fields must be top-level keys — never nest under EXPOSED_FRAME, PANELS, ROOM_LOCATION, NOTES
- exposed_frame_dimensions is always an array, never null — use [] if nothing found
  panel_shapes is always an array, never null — use [] if nothing found
- drawing_notes is always an array, never null — use [] if nothing found
- flags is always an array, never null — use [] if nothing found
- Use null for missing scalar fields
- Preserve all dimension strings exactly as written — do NOT convert fractions
- Return JSON only — no markdown, no explanation"""

# ── CLEANUP HELPERS ───────────────────────────────────────────────────────────
# Two cleanup functions for two different situations:
#
# _cleanup (delayed):
#   Used after successful jobs — waits 1 hour then deletes the temp folder
#   The 1 hour delay gives the frontend time to download results and reports
#   before the files are gone. Runs in a background daemon thread so it doesn't
#   block the server or keep the process alive after shutdown
#
# _cleanup_now (immediate):
#   Used after failed jobs — deletes the temp folder right away
#   No point keeping files from a job that already failed
#   Called before raising errors so we never leave orphaned folders on disk

def _cleanup(folder_path: str):
    # threading.Timer schedules shutil.rmtree to run after CLEANUP_DELAY_SEC seconds
    # daemon=True means this timer thread will be killed if the server shuts down
    # so we don't accidentally keep the process alive just for cleanup
    timer = threading.Timer(
        CLEANUP_DELAY_SEC,
        shutil.rmtree,
        args=[folder_path],
        kwargs={"ignore_errors": True}  # don't crash if folder is already gone
    )
    timer.daemon = True
    timer.start()  # non-blocking — returns immediately, cleanup happens later


def _cleanup_now(folder_path: str):
    # ignore_errors=True means this never raises an exception
    # even if the folder doesn't exist or we don't have permission
    shutil.rmtree(folder_path, ignore_errors=True)


# ── CLI ENTRY POINT ───────────────────────────────────────────────────────────
# This block only runs when you execute the file directly from the terminal
# It does NOT run when another file imports fastapi_app
#
# Usage:
#   python fastapi_app.py             → start server on port 8000
#   python fastapi_app.py --reload    → auto-restart when any .py file changes
#   python fastapi_app.py --port 9000 → run on a different port
#   python fastapi_app.py --host 127.0.0.1 → only accessible from localhost
if __name__ == "__main__":
    import argparse
    import uvicorn

    parser = argparse.ArgumentParser(description="Run the GFS FastAPI server")
    parser.add_argument("--host",   default="0.0.0.0",     help="Host to bind to (0.0.0.0 = all interfaces)")
    parser.add_argument("--port",   type=int, default=8000, help="Port to listen on")
    parser.add_argument("--reload", action="store_true",    help="Auto-reload on code changes (development only)")
    args = parser.parse_args()

    # Print a startup summary so the team can confirm settings at a glance
    print(f"\nStarting GFS API — Full Pipeline v2.0")
    print(f"  Docs:       http://localhost:{args.port}/docs")
    print(f"  Endpoint:   {AZURE_ENDPOINT}")
    print(f"  Model:      {AZURE_DEPLOYMENT}")
    print(f"  Temp dir:   {TEMP_BASE}")
    print(f"\n  Pipeline stages:")
    print(f"    1. PDF/ZIP upload and validation")
    print(f"    2. Convert PDF pages to PNG images    ← pdf_to_image.py")
    print(f"    3. Classify each page with Vision AI  ← page_classifier_ayush.py")
    print(f"    4. Extract COVER page data            ← cover prompt + Azure OpenAI")
    print(f"    5. Extract PLAN_VIEW page data        ← plan prompt + Azure OpenAI")
    print(f"    6. Return structured JSON + report_url")
    print("-" * 40)

    uvicorn.run("fastapi_app:app", host=args.host, port=args.port, reload=args.reload)
    