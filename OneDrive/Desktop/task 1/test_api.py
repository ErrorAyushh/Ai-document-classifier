# test_api.py
# ─────────────────────────────────────────────────────────────────────────────
# GFS Pipeline Test API — Completely separate from production FastAPI app.
# Run this independently to test the full pipeline end-to-end before demo.
#
# Usage:
#   uvicorn test_api:app --reload --port 8001
#
# Endpoints:
#   POST /upload   — accepts PDF or ZIP, runs full pipeline, returns detailed JSON
#   GET  /health   — confirms test server is alive
# ─────────────────────────────────────────────────────────────────────────────

import json
import os
import shutil
import sys
import tempfile
import traceback
import uuid
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl
from fastapi import FastAPI, File, UploadFile
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from decision_tree import build_job_data
from excel_filler import fill_workbook
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse

# ── Load .env before any project imports ─────────────────────────────────────
from dotenv import load_dotenv
load_dotenv()

# ── Project imports ───────────────────────────────────────────────────────────
# These import directly from your existing modules — nothing is rewritten.
# Adjust the import paths below if your project uses a package prefix
# (e.g. "from ai_engine.non_ai.pdf_processor import extract_pages").

try:

    from pdf_processor import extract_pages, render_page
    from page_classifier import classify_page
    from cover_extractor import extract_cover
    from plan_extractor import extract_plan
    from ocr_plan_extractor import extract_plan_ocr
    from job_aggregator import aggregate_pages
    from page_validator import validate_page
    from dim_converter import convert_dim_group
    from rules_engine import apply_all_rules
    from confidence_scorer import compute_confidence

    _IMPORT_ERROR = None
except ImportError as e:
    _IMPORT_ERROR = str(e)

# Vision client — separate try so we can still test non-AI pipeline if vision is unavailable
try:
    from vision_client import call_vision
    _VISION_AVAILABLE = True
except ImportError:
    _VISION_AVAILABLE = False

# COMBINED_EXTRACTION_PROMPT — used for SECTION / GLASS_DETAIL / SCANNED pages
COMBINED_EXTRACTION_PROMPT = (
    "You are a GFS shop drawing extraction engine. "
    "Identify the page type and extract all relevant fields. "
    "Return ONLY a valid JSON object with no markdown fences."
)
try:
    from extraction_prompts import COMBINED_EXTRACTION_PROMPT  # type: ignore
except ImportError:
    pass  # fallback string above is used


# ── FastAPI app ───────────────────────────────────────────────────────────────
app = FastAPI(
    title="GFS Pipeline Test API",
    description="Test the full GFS extraction pipeline end-to-end. Separate from production.",
    version="1.0.0-test",
)
app.mount("/static", StaticFiles(directory="static"), name="static")



@app.get("/")
def home():
    return FileResponse("static/index.html")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

ALLOWED_EXTENSIONS = {".pdf", ".zip"}
TEMP_BASE          = tempfile.gettempdir()


# ─────────────────────────────────────────────────────────────────────────────
# HEALTH CHECK
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/health")
def health():
    return {
        "status":           "ok",
        "import_error":     _IMPORT_ERROR,
        "vision_available": _VISION_AVAILABLE,
        "timestamp":        datetime.utcnow().isoformat(),
    }


# ─────────────────────────────────────────────────────────────────────────────
# UPLOAD ENDPOINT
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/upload")
async def upload(file: UploadFile = File(...)):
    """
    Accepts a PDF or ZIP file and runs the full pipeline with detailed logging.
    Returns a comprehensive JSON response showing every stage result.
    """

    # ── Guard: import check ───────────────────────────────────────────────────
    if _IMPORT_ERROR:
        return JSONResponse(
            status_code=500,
            content={
                "success": False,
                "error":   f"Import error — check module paths: {_IMPORT_ERROR}",
            },
        )

    # ── Setup temp folder ─────────────────────────────────────────────────────
    run_id      = str(uuid.uuid4())[:8]
    temp_folder = os.path.join(TEMP_BASE, f"gfs_test_{run_id}")
    os.makedirs(temp_folder, exist_ok=True)

    _log("UPLOAD", f"run_id={run_id}  filename={file.filename}")

    # ── Save uploaded file ────────────────────────────────────────────────────
    _, ext = os.path.splitext(file.filename or "")
    ext = ext.lower()

    if ext not in ALLOWED_EXTENSIONS:
        return JSONResponse(
            status_code=415,
            content={"success": False, "error": f"Unsupported file type: {ext}"},
        )

    saved_path = os.path.join(temp_folder, file.filename)
    with open(saved_path, "wb") as f:
        shutil.copyfileobj(file.file, f)

    _log("UPLOAD", f"saved to {saved_path}")

    # ── Find PDF path (handle ZIP) ────────────────────────────────────────────
    pdf_path = None
    if ext == ".zip":
        pdf_path = _extract_pdf_from_zip(saved_path, temp_folder)
        if not pdf_path:
            return JSONResponse(
                status_code=422,
                content={"success": False, "error": "No PDF found inside ZIP."},
            )
        _log("UPLOAD", f"ZIP extracted — PDF found: {pdf_path}")
    else:
        pdf_path = saved_path
        _log("UPLOAD", f"Direct PDF upload: {pdf_path}")

    # ── Run pipeline ──────────────────────────────────────────────────────────
    try:
        result = _run_test_pipeline(pdf_path, temp_folder, run_id)
    except Exception as exc:
        _log("ERROR", traceback.format_exc())
        return JSONResponse(
            status_code=500,
            content={
                "success": False,
                "run_id":  run_id,
                "error":   str(exc),
                "trace":   traceback.format_exc(),
            },
        )
    finally:
        # Clean up temp folder
        shutil.rmtree(temp_folder, ignore_errors=True)

    return JSONResponse(status_code=200, content=result)


# ─────────────────────────────────────────────────────────────────────────────
# CORE TEST PIPELINE
# ─────────────────────────────────────────────────────────────────────────────

def _run_test_pipeline(pdf_path: str, temp_folder: str, run_id: str) -> dict:
    """
    Runs every stage of the pipeline with detailed logging.
    Returns a comprehensive dict capturing every stage result.
    """

    report = {
        "success":          True,
        "run_id":           run_id,
        "pdf_path":         pdf_path,
        "timestamp":        datetime.utcnow().isoformat(),
        "pages_processed":  0,
        "vision_calls":     0,
        "token_summary":    {"prompt": 0, "completion": 0, "total": 0},
        "classification":   [],
        "page_results":     [],
        "aggregated_result": None,
        "rules_output":     None,
        "confidence":       None,
        "excel_path":       None,
        "errors":           [],
    }

    # ── STAGE 1: Extract text from PDF ───────────────────────────────────────
    _log("PDF_EXTRACT", f"Opening {pdf_path}")
    pages = extract_pages(pdf_path)
    _log("PDF_EXTRACT", f"Found {len(pages)} pages")

    # ── STAGE 2: Classify each page ───────────────────────────────────────────
    _log("CLASSIFIER", "Starting page classification")
    classifications = []
    for page in pages:
        pn   = page["page_number"]
        text = page.get("text", "")
        is_s = page.get("is_scanned", False)

        classification = classify_page(text, pn)
        classifications.append(classification)

        _log("CLASSIFIER",
             f"  Page {pn}: type={classification['page_type']}  "
             f"confidence={classification['confidence']}  "
             f"scanned={is_s}  text_len={len(text)}")

    report["classification"] = classifications

    # ── STAGE 3: Process each page through the extraction chain ───────────────
    all_page_results = []  # fed to job_aggregator
    page_detail_log  = []  # fed to API response

    for page, classification in zip(pages, classifications):
        pn        = page["page_number"]
        text      = page.get("text", "")
        is_scanned= page.get("is_scanned", False)
        page_type = classification["page_type"]

        _log("PIPELINE", f"─── Page {pn}: {page_type} ───")

        page_log = {
            "page_number":      pn,
            "page_type":        page_type,
            "is_scanned":       is_scanned,
            "non_ai_result":    None,
            "ocr_result":       None,
            "merged_non_ai_ocr": None,
            "targeted_prompt":  None,
            "vision_result":    None,
            "merged_vision":    None,
            "final_extraction": None,
            "method_used":      None,
            "vision_triggered": False,
            "stage_stopped_at": None,
        }

        # -- Skip OTHER pages completely ---------------------------------------
        if page_type == "OTHER":
            _log("PIPELINE", f"  Page {pn}: OTHER — skipped")
            page_log["stage_stopped_at"] = "SKIPPED_OTHER"
            page_detail_log.append(page_log)
            continue

        # -- SCANNED pages → always Vision ------------------------------------
        if page_type == "SCANNED" or is_scanned:
            _log("VISION", f"  Page {pn}: SCANNED — rendering and sending to Vision")
            image_path = render_page(pdf_path, pn, temp_folder)
            vision_res = _call_vision_safe(image_path, COMBINED_EXTRACTION_PROMPT, 2000)
            _add_tokens(report, vision_res)
            report["vision_calls"] += 1
            page_log["vision_triggered"] = True
            page_log["vision_result"]    = _safe_serialize(vision_res)

            if vision_res.get("ok"):
                extraction = vision_res["data"]
                validate_page(extraction)
                _normalise_dims(extraction)
                page_log["final_extraction"] = extraction
                page_log["method_used"]      = "vision_scanned"
                all_page_results.append({"page_number": pn, "extraction": extraction})
            else:
                _log("ERROR", f"  Page {pn}: Vision failed — {vision_res.get('error')}")
                report["errors"].append({"page": pn, "stage": "VISION_SCANNED", "error": vision_res.get("error")})

            page_log["stage_stopped_at"] = "VISION_SCANNED"
            page_detail_log.append(page_log)
            continue

        # -- COVER page -------------------------------------------------------
        if page_type == "COVER":
            extraction, page_log = _process_cover(
                pdf_path, pn, text, temp_folder, report, page_log
            )
            if extraction:
                all_page_results.append({"page_number": pn, "extraction": extraction})
            page_detail_log.append(page_log)
            continue

        # -- PLAN_VIEW page ---------------------------------------------------
        if page_type == "PLAN_VIEW":
            extraction, page_log = _process_plan(
                pdf_path, pn, text, temp_folder, report, page_log
            )
            if extraction:
                all_page_results.append({"page_number": pn, "extraction": extraction})
            page_detail_log.append(page_log)
            continue

        # -- SECTION / GLASS_DETAIL / MGDS_CATALOG → Vision ------------------
        if page_type in ("SECTION", "GLASS_DETAIL", "MGDS_CATALOG"):
            _log("VISION", f"  Page {pn}: {page_type} — sending to Vision")
            image_path = render_page(pdf_path, pn, temp_folder)
            vision_res = _call_vision_safe(image_path, COMBINED_EXTRACTION_PROMPT, 2000)
            _add_tokens(report, vision_res)
            report["vision_calls"] += 1
            page_log["vision_triggered"] = True
            page_log["vision_result"]    = _safe_serialize(vision_res)

            if vision_res.get("ok"):
                extraction = vision_res["data"]
                validate_page(extraction)
                _normalise_dims(extraction)
                page_log["final_extraction"] = extraction
                page_log["method_used"]      = f"vision_{page_type.lower()}"
                all_page_results.append({"page_number": pn, "extraction": extraction})
            else:
                _log("ERROR", f"  Page {pn}: Vision failed — {vision_res.get('error')}")
                report["errors"].append({"page": pn, "stage": f"VISION_{page_type}", "error": vision_res.get("error")})

            page_log["stage_stopped_at"] = f"VISION_{page_type}"
            page_detail_log.append(page_log)
            continue

    report["page_results"]    = page_detail_log
    report["pages_processed"] = len(all_page_results)

    # ── STAGE 4: Job Aggregator ───────────────────────────────────────────────
    _log("AGGREGATOR", f"Aggregating {len(all_page_results)} page results")
    if all_page_results:
        try:
            merged = aggregate_pages(all_page_results)
            _log("AGGREGATOR", f"  job_type={merged.get('job_type')}  "
                               f"warnings={merged.get('_merge_warnings', [])}")
            report["aggregated_result"] = _safe_serialize(merged)

            # ── STAGE 5: Rules Engine ─────────────────────────────────────────
            _log("RULES", "Running rules engine")
            cover_data = _build_cover_data_for_rules(merged)
            plan_data  = merged.get("plan_dims")
            rules_out  = apply_all_rules(cover_data, plan_data)
            report["rules_output"] = _safe_serialize(rules_out)
            _log("RULES", f"  series={rules_out.get('series')}  "
                          f"confidence={rules_out.get('confidence')}  "
                          f"flags={rules_out.get('flags', [])}")

            # ── STAGE 6: Confidence Scorer ────────────────────────────────────
            _log("CONFIDENCE", "Computing confidence score")
            conf = compute_confidence(rules_out)
            report["confidence"] = _safe_serialize(conf)
            _log("CONFIDENCE", f"  result={conf}")

            # ── STAGE 7: Excel generation ─────────────────────────────────────
            # ── STAGE 7: Real Excel Generation ──

            _log("EXCEL", "Building job_data")

            cover_result = {
                "project_header": {
                    "project_address": merged.get("project_address"),
                    "quote_number": "",
                    "revision": "",
                    "contractor": merged.get("approved_by_company"),
                    "drawing_date": merged.get("approval_date"),
                },

                "glass_specification": {
                    "glass_makeup": (
                        merged.get("glass_makeup")[0]
                        if merged.get("glass_makeup")
                        else None
                    ),
                    "expedited": merged.get("expedited", False),
                    "hst": False,
                    "back_paint": False,
                },

                "frame": {
                    "series": rules_out.get("series", "Series 2000"),
                    "frame_material": None,
                },

                "units": []
            }

            plan_result = {
                "unit_letter": merged.get("unit_letter"),
                "unit_qty": merged.get("unit_qty"),

                "shape": merged.get("shape", "RECTANGULAR"),

                "panel_count": merged.get("panel_count"),

                "drawing_notes": [],

                "exposed_frame_width": {
                    "raw": merged.get("plan_dims", {}).get(
                        "exposed_frame_width_raw"
                    ),
                    "decimal": merged.get("plan_dims", {}).get(
                        "exposed_frame_width_in"
                    ),
                }
                if merged.get("plan_dims")
                else None,

                "exposed_frame_length": {
                    "raw": merged.get("plan_dims", {}).get(
                        "exposed_frame_height_raw"
                    ),
                    "decimal": merged.get("plan_dims", {}).get(
                        "exposed_frame_height_in"
                    ),
                }
                if merged.get("plan_dims")
                else None,
            }

            job_data = build_job_data(
                cover_result,
                plan_result
            )

            _log("EXCEL", f"job_type={job_data['job_type']}")

            excel_path = os.path.join(
                TEMP_BASE,
                f"gfs_output_{run_id}.xlsx"
            )
            
            template_path = os.path.join(
            os.path.dirname(__file__),
            "templates",
            "gfs_pricing_template.xlsx"
                )

            _log("EXCEL", f"Template: {template_path}")
            _log("EXCEL", f"Exists: {os.path.exists(template_path)}")
            excel_result = fill_workbook(
                job_data=job_data,
                template_path=template_path,
                output_path=excel_path
)

            report["excel_path"] = excel_result["output_path"]

            _log(
                "EXCEL",
                f"Saved to {excel_result['output_path']}"
            )

        except Exception as exc:
            _log("AGGREGATOR_ERROR", traceback.format_exc())
            report["errors"].append({"stage": "AGGREGATOR", "error": str(exc)})
    else:
        _log("AGGREGATOR", "No page results to aggregate")

    # ── Final summary ─────────────────────────────────────────────────────────
    _log("RESPONSE", f"Pipeline complete — "
                     f"pages={report['pages_processed']}  "
                     f"vision_calls={report['vision_calls']}  "
                     f"errors={len(report['errors'])}")

    return report


# ─────────────────────────────────────────────────────────────────────────────
# COVER PAGE PROCESSING
# ─────────────────────────────────────────────────────────────────────────────

def _process_cover(pdf_path, pn, text, temp_folder, report, page_log):
    """Runs Non-AI cover extraction with Vision fallback if needed."""

    # Step 1: Non-AI
    _log("NON-AI", f"  Page {pn}: Running cover_extractor")
    non_ai = extract_cover(text, pn)
    page_log["non_ai_result"] = _safe_serialize(non_ai)
    _log("NON-AI", f"  Page {pn}: needs_vision={non_ai.get('needs_vision')}  "
                   f"address={non_ai.get('project_header', {}).get('project_address')}  "
                   f"glass_makeup={bool(non_ai.get('glass_specification', {}).get('glass_makeup'))}")

    if not non_ai.get("needs_vision"):
        # Non-AI succeeded
        extraction = _cover_to_aggregator(non_ai)
        page_log["final_extraction"] = extraction
        page_log["method_used"]      = "non_ai"
        page_log["stage_stopped_at"] = "NON_AI"
        _log("NON-AI", f"  Page {pn}: COVER done at Non-AI ✓")
        return extraction, page_log

    # Step 2: Vision fallback for cover
    _log("VISION", f"  Page {pn}: COVER needs_vision=True — calling Vision")
    image_path = render_page(pdf_path, pn, temp_folder)
    vision_res = _call_vision_safe(image_path, COMBINED_EXTRACTION_PROMPT, 2000)
    _add_tokens(report, vision_res)
    report["vision_calls"] += 1
    page_log["vision_triggered"] = True
    page_log["vision_result"]    = _safe_serialize(vision_res)

    if vision_res.get("ok"):
        extraction = vision_res["data"]
        validate_page(extraction)
        _normalise_dims(extraction)
        page_log["final_extraction"] = extraction
        page_log["method_used"]      = "vision_cover"
        page_log["stage_stopped_at"] = "VISION_COVER"
        _log("VISION", f"  Page {pn}: COVER Vision succeeded")
    else:
        # Fall back to best non-AI result
        extraction = _cover_to_aggregator(non_ai)
        page_log["final_extraction"] = extraction
        page_log["method_used"]      = "non_ai_fallback"
        page_log["stage_stopped_at"] = "NON_AI_FALLBACK"
        report["errors"].append({"page": pn, "stage": "VISION_COVER", "error": vision_res.get("error")})
        _log("ERROR", f"  Page {pn}: COVER Vision failed — using Non-AI fallback")

    return extraction, page_log


# ─────────────────────────────────────────────────────────────────────────────
# PLAN VIEW PAGE PROCESSING
# ─────────────────────────────────────────────────────────────────────────────

def _process_plan(pdf_path, pn, text, temp_folder, report, page_log):
    """
    Full plan page flow:
    Non-AI → OCR → Merge Non-AI+OCR → Targeted Prompt → Vision → Merge Vision → Map
    """

    # ── Step 1: Non-AI ────────────────────────────────────────────────────────
    _log("NON-AI", f"  Page {pn}: Running plan_extractor")
    non_ai = extract_plan(text, pn)
    page_log["non_ai_result"] = _safe_serialize(non_ai)

    _log("NON-AI",
         f"  Page {pn}: needs_vision={non_ai.get('needs_vision')}  "
         f"letter={non_ai.get('unit_letter')}  "
         f"qty={non_ai.get('unit_qty')}  "
         f"width={non_ai.get('exposed_frame_width')}  "
         f"length={non_ai.get('exposed_frame_length')}")

    if not non_ai.get("needs_vision"):
        extraction = _map_plan_to_aggregator(non_ai)
        page_log["final_extraction"] = extraction
        page_log["method_used"]      = "non_ai"
        page_log["stage_stopped_at"] = "NON_AI"
        _log("NON-AI", f"  Page {pn}: PLAN done at Non-AI ✓")
        return extraction, page_log

    # ── Step 2: OCR ───────────────────────────────────────────────────────────
    _log("OCR", f"  Page {pn}: Non-AI incomplete — running OCR")
    image_path = render_page(pdf_path, pn, temp_folder)
    ocr_result = extract_plan_ocr(image_path, pn)
    page_log["ocr_result"] = _safe_serialize(ocr_result)

    _log("OCR",
         f"  Page {pn}: needs_vision={ocr_result.get('needs_vision')}  "
         f"letter={ocr_result.get('unit_letter')}  "
         f"qty={ocr_result.get('unit_qty')}  "
         f"width={ocr_result.get('exposed_frame_width')}  "
         f"length={ocr_result.get('exposed_frame_length')}  "
         f"ocr_lines={ocr_result.get('ocr_line_count')}")

    if not ocr_result.get("needs_vision"):
        extraction = _map_plan_to_aggregator(ocr_result)
        page_log["final_extraction"] = extraction
        page_log["method_used"]      = "ocr"
        page_log["stage_stopped_at"] = "OCR"
        _log("OCR", f"  Page {pn}: PLAN done at OCR ✓")
        return extraction, page_log

    # ── Step 3: Merge Non-AI + OCR ────────────────────────────────────────────
    _log("MERGE_NON_AI_OCR", f"  Page {pn}: Both incomplete — merging Non-AI + OCR")
    merged_base = _merge_non_ai_and_ocr(non_ai, ocr_result)
    page_log["merged_non_ai_ocr"] = _safe_serialize(merged_base)

    _log("MERGE_NON_AI_OCR",
         f"  Page {pn}: after merge — "
         f"letter={merged_base.get('unit_letter')}  "
         f"qty={merged_base.get('unit_qty')}  "
         f"width={merged_base.get('exposed_frame_width')}  "
         f"length={merged_base.get('exposed_frame_length')}  "
         f"needs_vision={merged_base.get('needs_vision')}")

    if not merged_base.get("needs_vision"):
        extraction = _map_plan_to_aggregator(merged_base)
        page_log["final_extraction"] = extraction
        page_log["method_used"]      = "non_ai+ocr"
        page_log["stage_stopped_at"] = "MERGE_NON_AI_OCR"
        _log("MERGE_NON_AI_OCR", f"  Page {pn}: PLAN done after merge ✓")
        return extraction, page_log

    # ── Step 4: Build targeted prompt ─────────────────────────────────────────
    targeted_prompt = _build_targeted_prompt(merged_base)
    page_log["targeted_prompt"] = targeted_prompt

    _log("TARGETED_PROMPT",
         f"  Page {pn}: Generated targeted prompt ({len(targeted_prompt)} chars):")
    for line in targeted_prompt.splitlines():
        _log("TARGETED_PROMPT", f"    {line}")

    if not targeted_prompt:
        # Nothing left missing — use merged result
        extraction = _map_plan_to_aggregator(merged_base)
        page_log["final_extraction"] = extraction
        page_log["method_used"]      = "non_ai+ocr"
        page_log["stage_stopped_at"] = "MERGE_NON_AI_OCR"
        return extraction, page_log

    # ── Step 5: Vision — targeted call ────────────────────────────────────────
    _log("VISION", f"  Page {pn}: Calling Vision with targeted prompt")
    vision_res = _call_vision_safe(image_path, targeted_prompt, 500)
    _add_tokens(report, vision_res)
    report["vision_calls"] += 1
    page_log["vision_triggered"] = True
    page_log["vision_result"]    = _safe_serialize(vision_res)

    _log("VISION",
         f"  Page {pn}: Vision ok={vision_res.get('ok')}  "
         f"data={vision_res.get('data')}  "
         f"tokens={vision_res.get('token_usage', {})}")

    # ── Step 6: Merge Vision result ───────────────────────────────────────────
    if vision_res.get("ok") and vision_res.get("data"):
        merged_final = _merge_plan_results(merged_base, vision_res["data"])
        page_log["merged_vision"] = _safe_serialize(merged_final)

        _log("MERGE_VISION",
             f"  Page {pn}: after Vision merge — "
             f"letter={merged_final.get('unit_letter')}  "
             f"width={merged_final.get('exposed_frame_width')}  "
             f"length={merged_final.get('exposed_frame_length')}")

        extraction = _map_plan_to_aggregator(merged_final)
        page_log["method_used"]      = merged_final.get("method", "non_ai+vision")
        page_log["stage_stopped_at"] = "VISION"
    else:
        _log("ERROR", f"  Page {pn}: Vision failed — using merged Non-AI+OCR")
        report["errors"].append({"page": pn, "stage": "VISION_PLAN", "error": vision_res.get("error")})
        extraction = _map_plan_to_aggregator(merged_base)
        page_log["method_used"]      = "non_ai+ocr_fallback"
        page_log["stage_stopped_at"] = "OCR_FALLBACK"

    page_log["final_extraction"] = extraction
    _log("MAPPING", f"  Page {pn}: mapped to aggregator schema — plan_dims={extraction.get('plan_dims')}")
    return extraction, page_log


# ─────────────────────────────────────────────────────────────────────────────
# MERGE HELPERS (mirrors orchestrator.py logic exactly)
# ─────────────────────────────────────────────────────────────────────────────

def _merge_non_ai_and_ocr(non_ai_result: dict, ocr_result: dict) -> dict:
    """
    Merge Non-AI and OCR results.
    CASE 1: Non-AI has value, OCR missing  → keep Non-AI
    CASE 2: Non-AI missing, OCR has value  → use OCR
    CASE 3: Both have values               → OCR wins
    """
    merged = {}
    merged["page_number"] = ocr_result.get("page_number") or non_ai_result.get("page_number")
    merged["method"]      = "non_ai+ocr"

    # unit_letter
    ocr_l = ocr_result.get("unit_letter")
    nai_l = non_ai_result.get("unit_letter")
    merged["unit_letter"] = ocr_l if ocr_l else nai_l

    # unit_qty
    ocr_q = ocr_result.get("unit_qty")
    nai_q = non_ai_result.get("unit_qty")
    merged["unit_qty"] = ocr_q if ocr_q is not None else nai_q

    # exposed_frame_width
    ocr_w = ocr_result.get("exposed_frame_width")
    nai_w = non_ai_result.get("exposed_frame_width")
    merged["exposed_frame_width"] = ocr_w if ocr_w else nai_w

    # exposed_frame_length
    ocr_ln = ocr_result.get("exposed_frame_length")
    nai_ln = non_ai_result.get("exposed_frame_length")
    merged["exposed_frame_length"] = ocr_ln if ocr_ln else nai_ln

    # shape — prefer non-AI if it detected something specific
    nai_shape = non_ai_result.get("shape", "RECTANGULAR")
    ocr_shape = ocr_result.get("shape", "RECTANGULAR")
    merged["shape"] = nai_shape if nai_shape != "RECTANGULAR" else ocr_shape

    # panel_count
    ocr_pc = ocr_result.get("panel_count")
    nai_pc = non_ai_result.get("panel_count")
    merged["panel_count"] = ocr_pc if ocr_pc is not None else nai_pc

    # drawing_notes — union
    ocr_notes = ocr_result.get("drawing_notes") or []
    nai_notes = non_ai_result.get("drawing_notes") or []
    seen, combined = set(), []
    for note in (ocr_notes + nai_notes):
        if note and note not in seen:
            seen.add(note)
            combined.append(note)
    merged["drawing_notes"] = combined

    # Recalculate needs_vision
    has_width  = merged.get("exposed_frame_width")  is not None
    has_length = merged.get("exposed_frame_length") is not None
    has_letter = merged.get("unit_letter")          is not None
    merged["needs_vision"] = not (has_width and has_length and has_letter)

    return merged


def _build_targeted_prompt(base_result: dict) -> str:
    """Build Vision prompt asking only for missing fields."""
    missing_fields   = []
    missing_sections = []

    if not base_result.get("unit_letter"):
        missing_fields.append(
            '- "unit_letter": single uppercase letter from the page title '
            '(e.g. "PLAN - UNIT A" → "A"). Return null if not found.'
        )
        missing_sections.append('  "unit_letter": null')

    if base_result.get("unit_qty") is None:
        missing_fields.append(
            '- "unit_qty": integer quantity from "QTY X" in the page title. '
            'Return null if not found.'
        )
        missing_sections.append('  "unit_qty": null')

    if not base_result.get("exposed_frame_width"):
        missing_fields.append(
            '- "exposed_frame_width_raw": the SMALLER dimension labelled '
            '"(EXPOSED FRAME)" or "(EXP. FRAME)". '
            'Preserve raw string exactly (e.g. "47-5/8\\"", "57\\""). '
            'Return null if not found.'
        )
        missing_sections.append('  "exposed_frame_width_raw": null')

    if not base_result.get("exposed_frame_length"):
        missing_fields.append(
            '- "exposed_frame_length_raw": the LARGER dimension labelled '
            '"(EXPOSED FRAME)" or "(EXP. FRAME)". '
            'Preserve raw string exactly. Return null if not found.'
        )
        missing_sections.append('  "exposed_frame_length_raw": null')

    if not missing_fields:
        return ""

    fields_block = "\n\n".join(missing_fields)
    schema_block = "{\n" + ",\n".join(missing_sections) + "\n}"

    return (
        "You are extracting specific missing fields from a GFS PLAN VIEW drawing.\n\n"
        "Extract ONLY the following fields. Do not extract anything else.\n\n"
        f"FIELDS TO EXTRACT:\n{fields_block}\n\n"
        "RULES:\n"
        "- Only extract dimensions labelled EXPOSED FRAME or EXP. FRAME\n"
        "- Preserve dimension strings exactly as written\n"
        "- Return null for anything you cannot find with confidence\n\n"
        f"Return ONLY this JSON — no markdown, no explanation:\n{schema_block}"
    )


def _merge_plan_results(base_result: dict, vision_data: dict) -> dict:
    """Merge targeted Vision result into base (Non-AI+OCR merged) result."""
    merged = dict(base_result)
    merged.pop("needs_vision", None)
    merged["method"] = "non_ai+vision"

    v_letter = vision_data.get("unit_letter")
    if v_letter:
        merged["unit_letter"] = str(v_letter).strip().upper()

    v_qty = vision_data.get("unit_qty")
    if v_qty is not None:
        try:
            merged["unit_qty"] = int(v_qty)
        except (TypeError, ValueError):
            pass

    v_width_raw = vision_data.get("exposed_frame_width_raw")
    if v_width_raw:
        dec = _raw_to_decimal(v_width_raw)
        if dec is not None:
            merged["exposed_frame_width"] = {"raw": str(v_width_raw).strip(), "decimal": dec}

    v_length_raw = vision_data.get("exposed_frame_length_raw")
    if v_length_raw:
        dec = _raw_to_decimal(v_length_raw)
        if dec is not None:
            merged["exposed_frame_length"] = {"raw": str(v_length_raw).strip(), "decimal": dec}

    return merged


def _map_plan_to_aggregator(result: dict) -> dict:
    """Convert flat plan result to aggregator-compatible schema."""
    width  = result.get("exposed_frame_width")
    length = result.get("exposed_frame_length")

    plan_dims = None
    if width or length:
        plan_dims = {
            "exposed_frame_width_raw":  width["raw"]     if width  else None,
            "exposed_frame_width_in":   width["decimal"] if width  else None,
            "exposed_frame_height_raw": length["raw"]    if length else None,
            "exposed_frame_height_in":  length["decimal"] if length else None,
            "_conversion_warnings":     [],
        }

    return {
        "page_type":         "PLAN_VIEW",
        "sheet_id":          None,
        "unit_letter":       result.get("unit_letter"),
        "unit_qty":          result.get("unit_qty"),
        "shape":             result.get("shape", "RECTANGULAR"),
        "panel_count":       result.get("panel_count"),
        "plan_dims":         plan_dims,
        "section_dims":      None,
        "glass_detail_dims": None,
        "method":            result.get("method", "non_ai"),
    }


def _cover_to_aggregator(non_ai_cover: dict) -> dict:
    """Convert non-AI cover result to aggregator-compatible schema."""
    ph = non_ai_cover.get("project_header", {}) or {}
    gs = non_ai_cover.get("glass_specification", {}) or {}
    fr = non_ai_cover.get("frame", {}) or {}

    return {
        "page_type":            "COVER",
        "sheet_id":             None,
        "project_address":      ph.get("project_address"),
        "approval_date":        ph.get("drawing_date"),
        "approved_by_name":     None,
        "approved_by_title":    None,
        "approved_by_company":  ph.get("contractor"),
        "glass_makeup":         [gs["glass_makeup"]] if gs.get("glass_makeup") else None,
        "expedited":            gs.get("expedited", False),
        "job_type":             "STANDARD_FRAMED_UNIT",
        "plan_dims":            None,
        "section_dims":         None,
        "glass_detail_dims":    None,
        "method":               non_ai_cover.get("method", "non_ai"),
    }


# ─────────────────────────────────────────────────────────────────────────────
# DIMENSION CONVERTER (inline — avoids circular imports)
# ─────────────────────────────────────────────────────────────────────────────

def _raw_to_decimal(raw: str):
    """Convert raw GFS dimension string to decimal inches."""
    import re as _re
    try:
        s = str(raw).strip().replace('"', '').replace('\u2019', "'")
        m = _re.match(r"^(\d+)'\s*(\d+)[\-\s](\d+)/(\d+)$", s)
        if m:
            return int(m.group(1))*12 + int(m.group(2)) + int(m.group(3))/int(m.group(4))
        m = _re.match(r"^(\d+)'\s*(\d+)$", s)
        if m:
            return int(m.group(1))*12 + int(m.group(2))
        m = _re.match(r"^(\d+)[\-\s](\d+)/(\d+)$", s)
        if m:
            return int(m.group(1)) + int(m.group(2))/int(m.group(3))
        m = _re.match(r"^(\d+(?:\.\d+)?)$", s)
        if m:
            return float(m.group(1))
        return None
    except (ValueError, ZeroDivisionError):
        return None


# ─────────────────────────────────────────────────────────────────────────────
# DIM NORMALISER
# ─────────────────────────────────────────────────────────────────────────────

def _normalise_dims(extraction: dict) -> dict:
    """Run convert_dim_group on known dim fields."""
    for field in ("plan_dims", "section_dims", "glass_detail_dims"):
        if extraction.get(field) is not None:
            try:
                extraction[field] = convert_dim_group(extraction[field])
            except Exception:
                pass
    return extraction


# ─────────────────────────────────────────────────────────────────────────────
# RULES ENGINE ADAPTER
# ─────────────────────────────────────────────────────────────────────────────

def _build_cover_data_for_rules(merged: dict) -> dict:
    """Build cover_data dict in the shape rules_engine.apply_all_rules() expects."""
    return {
        "project_header": {
            "project_address": merged.get("project_address"),
            "quote_number":    None,
            "revision":        None,
            "contractor":      merged.get("approved_by_company"),
            "drawing_date":    merged.get("approval_date"),
        },
        "glass_specification": {
            "glass_makeup": merged.get("glass_makeup"),
            "expedited":    merged.get("expedited") or False,
            "hst":          False,
            "back_paint":   False,
        },
        "frame": {
            "series":         "Series 2000",
            "frame_material": None,
        },
        "units": [],
    }


# ─────────────────────────────────────────────────────────────────────────────
# VISION SAFE WRAPPER
# ─────────────────────────────────────────────────────────────────────────────

def _call_vision_safe(image_path: str, prompt: str, max_tokens: int) -> dict:
    """Call Vision with graceful fallback if vision client unavailable."""
    if not _VISION_AVAILABLE:
        return {
            "ok":          False,
            "error":       "vision_client not available",
            "token_usage": {},
        }
    try:
        return call_vision(image_path, prompt, max_tokens)
    except Exception as exc:
        return {
            "ok":          False,
            "error":       str(exc),
            "token_usage": {},
        }


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL GENERATOR
# ─────────────────────────────────────────────────────────────────────────────

def _generate_excel(merged: dict, rules_out: dict, confidence, output_path: str):
    import os
    import openpyxl

    template_path = os.path.join(
        os.path.dirname(__file__),
        "templates",
        "gfs_pricing_template.xlsx"
    )

    wb = openpyxl.load_workbook(template_path)

    # =========================
    # Pricing Worksheet
    # =========================
    ws = wb["Pricing Worksheet"]

    project_address = merged.get("project_address") or ""
    contractor = merged.get("approved_by_company") or ""
    approval_date = merged.get("approval_date") or ""
    job_type = merged.get("job_type") or ""

    glass_makeup = merged.get("glass_makeup")
    if isinstance(glass_makeup, list):
        glass_makeup = ", ".join(str(x) for x in glass_makeup)
    glass_makeup = glass_makeup or ""

    # Populate header fields
    ws["E2"] = project_address
    ws["B3"] = contractor
    ws["B4"] = job_type

    # =========================
    # AI Output Sheet
    # =========================
    if "AI_Output" in wb.sheetnames:
        del wb["AI_Output"]

    ai_ws = wb.create_sheet("AI_Output")

    ai_ws["A1"] = "Field"
    ai_ws["B1"] = "Value"

    row = 2

    entries = [
        ("Project Address", project_address),
        ("Contractor", contractor),
        ("Approval Date", approval_date),
        ("Job Type", job_type),
        ("Glass Makeup", glass_makeup),
        ("Series", rules_out.get("series") if rules_out else ""),
        ("Confidence", str(confidence)),
    ]

    for key, value in entries:
        ai_ws.cell(row=row, column=1).value = key
        ai_ws.cell(row=row, column=2).value = value
        row += 1

    # =========================
    # Plan Dimensions
    # =========================
    plan = merged.get("plan_dims") or {}

    row += 2
    ai_ws.cell(row=row, column=1).value = "Plan Dimensions"
    row += 1

    for k, v in plan.items():
        ai_ws.cell(row=row, column=1).value = k
        ai_ws.cell(row=row, column=2).value = str(v)
        row += 1

    # =========================
    # Rules Output
    # =========================
    if rules_out:
        row += 2
        ai_ws.cell(row=row, column=1).value = "Rules Output"
        row += 1

        for k, v in rules_out.items():
            ai_ws.cell(row=row, column=1).value = k
            ai_ws.cell(row=row, column=2).value = str(v)
            row += 1

    wb.save(output_path)


# ─────────────────────────────────────────────────────────────────────────────
# ZIP HANDLER
# ─────────────────────────────────────────────────────────────────────────────

def _extract_pdf_from_zip(zip_path: str, dest_folder: str):
    """
    Extract ZIP contents and return the PDF from the Shop Drawings folder.
    Falls back to first PDF if no Shop Drawings PDF exists.
    """

    try:
        with zipfile.ZipFile(zip_path, "r") as z:

            # Extract all files safely
            for member in z.infolist():

                # Clean each folder/file name for Windows
                clean_parts = []

                for part in Path(member.filename).parts:
                    for ch in ':*?"<>|':
                        part = part.replace(ch, "-")
                    clean_parts.append(part)

                target = os.path.join(dest_folder, *clean_parts)

                if member.is_dir():
                    os.makedirs(target, exist_ok=True)
                    continue

                os.makedirs(os.path.dirname(target), exist_ok=True)

                with z.open(member) as src, open(target, "wb") as dst:
                    shutil.copyfileobj(src, dst)

        # --------------------------------------------------
        # Find PDFs
        # --------------------------------------------------
        all_pdfs = []
        shop_pdfs = []

        for root, _, files in os.walk(dest_folder):
            for fname in files:

                if not fname.lower().endswith(".pdf"):
                    continue

                full_path = os.path.join(root, fname)

                all_pdfs.append(full_path)

                root_lower = root.lower()

                if (
                    "shop drawing" in root_lower
                    or "shop drawings" in root_lower
                    or "shop_drawing" in root_lower
                    or "shop_drawings" in root_lower
                ):
                    shop_pdfs.append(full_path)

        print("\n========== PDF SEARCH ==========")

        print("\nSHOP DRAWING PDFS:")
        for pdf in shop_pdfs:
            print(pdf)

        print("\nALL PDFS:")
        for pdf in all_pdfs:
            print(pdf)

        print("================================\n")

        # --------------------------------------------------
        # Priority 1: Shop Drawings PDF
        # --------------------------------------------------
        if shop_pdfs:
            selected = max(shop_pdfs, key=os.path.getsize)

            _log(
                "ZIP",
                f"Selected Shop Drawing PDF: {os.path.basename(selected)}"
            )

            return selected

        # --------------------------------------------------
        # Priority 2: Largest PDF in ZIP
        # --------------------------------------------------
        if all_pdfs:
            selected = max(all_pdfs, key=os.path.getsize)

            _log(
                "ZIP",
                f"No Shop Drawings folder found. Using largest PDF: {os.path.basename(selected)}"
            )

            return selected

    except Exception as exc:
        _log("ZIP_ERROR", str(exc))

    return None


def _sanitize(filename: str):
    illegal = ':*?"<>|'
    for ch in illegal:
        filename = filename.replace(ch, "-")
    return filename

# ─────────────────────────────────────────────────────────────────────────────
# TOKEN ACCUMULATOR
# ─────────────────────────────────────────────────────────────────────────────

def _add_tokens(report: dict, vision_res: dict):
    """Accumulate token usage from a Vision call into the report summary."""
    usage = vision_res.get("token_usage", {}) or {}
    report["token_summary"]["prompt"]     += usage.get("prompt_tokens",     0) or 0
    report["token_summary"]["completion"] += usage.get("completion_tokens", 0) or 0
    report["token_summary"]["total"]      += usage.get("total_tokens",      0) or 0


# ─────────────────────────────────────────────────────────────────────────────
# LOGGING HELPER
# ─────────────────────────────────────────────────────────────────────────────

def _log(stage: str, message: str):
    """Print a formatted log line to terminal."""
    print(f"[{stage:<20}] {message}", flush=True)


# ─────────────────────────────────────────────────────────────────────────────
# JSON SERIALIZER (handles non-serializable types)
# ─────────────────────────────────────────────────────────────────────────────

def _safe_serialize(obj):
    """Convert any object to a JSON-serializable form."""
    if obj is None:
        return None
    try:
        json.dumps(obj)
        return obj
    except (TypeError, ValueError):
        return str(obj)


# ─────────────────────────────────────────────────────────────────────────────
# CLI ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("test_api:app", host="0.0.0.0", port=8001, reload=True)